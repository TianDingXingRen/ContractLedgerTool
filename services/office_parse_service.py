"""Resource-bounded workers for parsing or generating Office documents."""

from __future__ import annotations

from openpyxl import load_workbook

from services.isolated_process import run_isolated_worker
from utils.contract_preview import build_preview_model
from utils.field_utils import detect_markers
from utils.generation_utils import generate_docx_document
from utils.security import MAX_TEMPLATE_FIELDS


OFFICE_WORKER_TIMEOUT_SECONDS = 60
OFFICE_WORKER_MEMORY_MB = 1536


def _detect_markers_worker(path, result_queue):
    result_queue.put(('ok', detect_markers(path)))


def detect_markers_isolated(path):
    fields = run_isolated_worker(
        _detect_markers_worker,
        (path,),
        timeout=OFFICE_WORKER_TIMEOUT_SECONDS,
        label='DOCX 占位符解析',
        memory_limit_mb=OFFICE_WORKER_MEMORY_MB,
    )
    if len(fields) > MAX_TEMPLATE_FIELDS:
        raise ValueError(f'模板字段数量不能超过 {MAX_TEMPLATE_FIELDS}')
    return fields


def _preview_worker(path, fields, result_queue):
    result_queue.put(('ok', build_preview_model(path, fields)))


def build_preview_model_isolated(path, fields):
    return run_isolated_worker(
        _preview_worker,
        (path, fields),
        timeout=OFFICE_WORKER_TIMEOUT_SECONDS,
        label='DOCX 预览解析',
        memory_limit_mb=OFFICE_WORKER_MEMORY_MB,
    )


def _generate_docx_worker(
    template_data,
    fields,
    field_values,
    source_docx,
    output_path,
    result_queue,
):
    result_queue.put(('ok', generate_docx_document(
        template_data,
        fields,
        field_values,
        source_docx,
        output_path,
    )))


def generate_docx_isolated(
    template_data,
    fields,
    field_values,
    source_docx,
    output_path,
):
    return run_isolated_worker(
        _generate_docx_worker,
        (template_data, fields, field_values, source_docx, output_path),
        timeout=OFFICE_WORKER_TIMEOUT_SECONDS,
        label='DOCX 文档生成',
        memory_limit_mb=OFFICE_WORKER_MEMORY_MB,
    )


def _extract_excel_rows_worker(path, max_rows, max_columns, result_queue):
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if sheet.max_row > max_rows:
            raise ValueError(f'Excel 行数不能超过 {max_rows}')
        if sheet.max_column > max_columns:
            raise ValueError(f'Excel 列数不能超过 {max_columns}')
        rows = list(sheet.iter_rows(
            min_row=1,
            max_row=max(1, sheet.max_row),
            max_col=max(1, sheet.max_column),
            values_only=True,
        ))
        result_queue.put(('ok', rows))
    finally:
        workbook.close()


def extract_excel_rows_isolated(path, *, max_rows, max_columns):
    return run_isolated_worker(
        _extract_excel_rows_worker,
        (path, int(max_rows), int(max_columns)),
        timeout=OFFICE_WORKER_TIMEOUT_SECONDS,
        label='Excel 明细解析',
        memory_limit_mb=OFFICE_WORKER_MEMORY_MB,
    )
