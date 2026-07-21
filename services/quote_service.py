"""Standard quotation workbook generation and import."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import procurement_store
from services import procurement_file_service
from utils.logger import get_logger
from utils.money import SQLITE_MAX_INTEGER, to_minor
from utils.security import safe_spreadsheet_value, validate_office_archive


FORMAT_VERSION = '1.0'
PARSER_VERSION = '1.0'
MAX_QUOTE_PDF_BYTES = 25 * 1024 * 1024
PDF_HEADER = b'%PDF-'

_HEADER_FILL = PatternFill('solid', fgColor='1D4ED8')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_INPUT_FILL = PatternFill('solid', fgColor='FFF7D6')
_LOCKED_FILL = PatternFill('solid', fgColor='E5E7EB')


def _decimal(value, label, errors, row=None, allow_zero=True):
    try:
        result = Decimal(str(value if value is not None else '').strip())
    except InvalidOperation:
        errors.append(f'{label}{f"（第 {row} 行）" if row else ""}格式无效')
        return None
    if not result.is_finite():
        errors.append(f'{label}{f"（第 {row} 行）" if row else ""}必须是有限数值')
        return None
    if result < 0 or (not allow_zero and result == 0):
        errors.append(f'{label}{f"（第 {row} 行）" if row else ""}必须大于 0')
        return None
    return result


def _money_minor(value):
    return to_minor(value, allow_none=False)


def _date_text(value):
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    return str(value or '').strip()


def generate_quote_template(project_id, supplier_id):
    project = procurement_store.get_project(project_id)
    supplier = procurement_store.get_project_supplier(supplier_id)
    items = procurement_store.list_project_items(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    if not supplier or supplier['project_id'] != project_id:
        raise ValueError('候选供应商不存在')
    if not items:
        raise ValueError('请先录入采购明细')

    workbook = Workbook()
    info = workbook.active
    info.title = '报价信息'
    info.column_dimensions['A'].width = 22
    info.column_dimensions['B'].width = 42
    info.append(['字段', '请供应商填写'])
    for cell in info[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    info_rows = [
        ('项目编号', project['project_no']),
        ('项目名称', project['project_name']),
        ('供应商名称', supplier['supplier_name']),
        ('报价轮次', 1),
        ('报价日期', ''),
        ('报价有效期', ''),
        ('税率(%)', 13),
        ('价格口径', '含税'),
        ('整体交付周期', ''),
        ('付款条件', ''),
        ('质保期', ''),
        ('包装运输', ''),
        ('整体技术偏离', ''),
        ('整体商务偏离', ''),
        ('报价总额', ''),
    ]
    for label, value in info_rows:
        info.append([label, safe_spreadsheet_value(value)])
        info.cell(info.max_row, 2).fill = _INPUT_FILL if label not in {'项目编号', '项目名称', '供应商名称'} else _LOCKED_FILL
    price_basis = DataValidation(type='list', formula1='"含税,不含税"', allow_blank=False)
    info.add_data_validation(price_basis)
    price_basis.add(info['B9'])

    detail = workbook.create_sheet('报价明细')
    headers = [
        '项目明细ID', '行号', '物资名称', '规格型号', '图号/代号', '数量', '单位',
        '单价', '金额', '分项交期', '技术偏离', '商务偏离', '备注',
    ]
    detail.append(headers)
    for cell in detail[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    widths = [14, 8, 24, 18, 16, 12, 10, 14, 16, 16, 24, 24, 24]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[detail.cell(1, index).column_letter].width = width
    for row_index, item in enumerate(items, start=2):
        detail.append([
            item['id'], item['line_no'], safe_spreadsheet_value(item['item_name']),
            safe_spreadsheet_value(item.get('spec_model') or ''),
            safe_spreadsheet_value(item.get('drawing_no') or ''),
            Decimal(item['quantity_text']), safe_spreadsheet_value(item['unit']), '',
            f'=ROUND(F{row_index}*H{row_index},2)',
            safe_spreadsheet_value(item.get('required_delivery_date') or ''),
            '', '', '',
        ])
        for col in range(1, 8):
            detail.cell(row_index, col).fill = _LOCKED_FILL
        for col in range(8, 14):
            detail.cell(row_index, col).fill = _INPUT_FILL
        detail.cell(row_index, 8).number_format = '0.00'
        detail.cell(row_index, 9).number_format = '0.00'
    detail.column_dimensions['A'].hidden = True
    detail.freeze_panes = 'A2'
    detail.auto_filter.ref = detail.dimensions

    readme = workbook.create_sheet('填报说明', 0)
    readme.column_dimensions['A'].width = 100
    readme.append(['供应商标准报价模板'])
    readme['A1'].font = Font(size=16, bold=True)
    instructions = [
        '1. 请勿删除或新增“报价明细”行，不得修改灰色字段。',
        '2. 黄色字段为供应商填写项；单价和总额统一按“价格口径”填写。',
        '3. 金额公式仅供填报核对，系统导入时会重新计算。',
        '4. 如存在技术或商务偏离，请在对应列明确填写。',
        f'5. 模板格式版本：{FORMAT_VERSION}',
    ]
    for line in instructions:
        readme.append([line])

    meta = workbook.create_sheet('_meta')
    meta.sheet_state = 'hidden'
    for key, value in [
        ('format_version', FORMAT_VERSION), ('project_id', project_id),
        ('project_no', project['project_no']), ('supplier_id', supplier_id),
    ]:
        meta.append([key, safe_spreadsheet_value(value)])

    filename = f"{project['project_no']}_{supplier['supplier_name']}_标准报价模板.xlsx"
    try:
        path = procurement_file_service.save_generated(
            project, 'quote_template', filename, workbook.save,
        )
    finally:
        workbook.close()
    if project['status'] == 'draft':
        procurement_store.transition_project_status(project_id, 'documents_ready', '已生成标准报价模板')
    return str(path)


def _sheet_pairs(sheet):
    values = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = str(row[0] or '').strip()
        if key:
            values[key] = row[1]
    return values


def parse_standard_quote(path, project_id, supplier_id, expected_round):
    errors = []
    warnings = []
    project = procurement_store.get_project(project_id)
    supplier = procurement_store.get_project_supplier(supplier_id)
    expected_items = {item['id']: item for item in procurement_store.list_project_items(project_id)}
    if not project or not supplier or supplier['project_id'] != project_id:
        return {}, ['采购项目或候选供应商不存在'], []
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        return {}, [f'Excel 文件无法读取：{exc}'], []
    try:
        return _parse_standard_quote_workbook(
            workbook,
            path,
            project_id,
            supplier_id,
            expected_round,
            project,
            supplier,
            expected_items,
            errors,
            warnings,
        )
    finally:
        workbook.close()


def _parse_standard_quote_workbook(
    workbook,
    path,
    project_id,
    supplier_id,
    expected_round,
    project,
    supplier,
    expected_items,
    errors,
    warnings,
):
    required_sheets = {'_meta', '报价信息', '报价明细'}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        return {}, [f'缺少工作表：{"、".join(sorted(missing))}'], []

    meta = {str(row[0].value): row[1].value for row in workbook['_meta'].iter_rows(min_row=1, max_col=2) if row[0].value}
    if str(meta.get('format_version') or '') != FORMAT_VERSION:
        errors.append(f'模板版本不匹配，要求 {FORMAT_VERSION}')
    if str(meta.get('project_id')) != str(project_id) or str(meta.get('project_no')) != project['project_no']:
        errors.append('报价模板不属于当前采购项目')
    if str(meta.get('supplier_id')) != str(supplier_id):
        errors.append('报价模板供应商与当前选择不一致')

    info = _sheet_pairs(workbook['报价信息'])
    if str(info.get('供应商名称') or '').strip() != supplier['supplier_name']:
        errors.append('报价信息中的供应商名称不匹配')
    try:
        quote_round = int(info.get('报价轮次') or expected_round)
    except (TypeError, ValueError):
        quote_round = expected_round
        errors.append('报价轮次必须为整数')
    if quote_round != int(expected_round):
        errors.append('报价文件轮次与本次导入轮次不一致')

    tax_raw = info.get('税率(%)')
    tax_rate_bps = None
    if tax_raw not in (None, ''):
        tax = _decimal(tax_raw, '税率', errors)
        if tax is not None:
            if tax > 100:
                errors.append('税率不能超过 100%')
            else:
                tax_rate_bps = int((tax * 100).quantize(Decimal('1')))
    price_basis = 'tax_inclusive' if str(info.get('价格口径') or '含税').strip() == '含税' else 'tax_exclusive'

    parsed_items = []
    seen = set()
    detail = workbook['报价明细']
    for excel_row, row in enumerate(detail.iter_rows(min_row=2, values_only=False), start=2):
        if all(cell.value in (None, '') for cell in row):
            continue
        try:
            item_id = int(row[0].value)
        except (TypeError, ValueError):
            errors.append(f'第 {excel_row} 行项目明细 ID 无效')
            continue
        expected = expected_items.get(item_id)
        if not expected:
            errors.append(f'第 {excel_row} 行项目明细不属于当前项目')
            continue
        if item_id in seen:
            errors.append(f'第 {excel_row} 行项目明细重复')
            continue
        seen.add(item_id)
        quantity = _decimal(row[5].value, '数量', errors, excel_row, allow_zero=False)
        unit_price = _decimal(row[7].value, '单价', errors, excel_row)
        if quantity is None or unit_price is None:
            continue
        if format(quantity.normalize(), 'f') != expected['quantity_text']:
            warnings.append(f'第 {excel_row} 行数量与项目需求不一致')
        if str(row[6].value or '').strip() != expected['unit']:
            warnings.append(f'第 {excel_row} 行单位与项目需求不一致')
        if unit_price == 0:
            warnings.append(f'第 {excel_row} 行单价为 0')
        amount = quantity * unit_price
        parsed_items.append({
            'project_item_id': item_id,
            'line_no': expected['line_no'],
            'item_name': expected['item_name'],
            'spec_model': expected.get('spec_model') or '',
            'drawing_no': expected.get('drawing_no') or '',
            'quantity_text': format(quantity.normalize(), 'f'),
            'unit': str(row[6].value or expected['unit']).strip(),
            'unit_price_minor': _money_minor(unit_price),
            'amount_minor': _money_minor(amount),
            'delivery_period': str(row[9].value or '').strip(),
            'technical_deviation': str(row[10].value or '').strip(),
            'commercial_deviation': str(row[11].value or '').strip(),
            'remark': str(row[12].value or '').strip(),
        })
    missing_items = sorted(set(expected_items) - seen)
    if missing_items:
        warnings.append(f'缺少 {len(missing_items)} 项项目明细')
    if not parsed_items:
        errors.append('报价明细为空')
    computed_total = sum((item['amount_minor'] for item in parsed_items), 0)
    stated_total = info.get('报价总额')
    if stated_total not in (None, ''):
        total_decimal = _decimal(stated_total, '报价总额', errors)
        if total_decimal is not None and _money_minor(total_decimal) != computed_total:
            errors.append('报价总额与明细重新计算结果不一致')

    payload = {
        'header': {
            'quote_round': quote_round,
            'quote_date': _date_text(info.get('报价日期')),
            'quote_valid_until': _date_text(info.get('报价有效期')),
            'total_amount_minor': computed_total,
            'currency': 'CNY',
            'tax_rate_bps': tax_rate_bps,
            'price_basis': price_basis,
            'delivery_period': str(info.get('整体交付周期') or '').strip(),
            'payment_terms': str(info.get('付款条件') or '').strip(),
            'warranty_period': str(info.get('质保期') or '').strip(),
            'package_transport': str(info.get('包装运输') or '').strip(),
            'technical_deviation': str(info.get('整体技术偏离') or '').strip(),
            'commercial_deviation': str(info.get('整体商务偏离') or '').strip(),
        },
        'items': parsed_items,
        'missing_project_item_ids': missing_items,
        'size_bytes': Path(path).stat().st_size,
    }
    return payload, errors, warnings


def create_import_job(project_id, supplier_id, quote_round, file_storage):
    project = procurement_store.get_project(project_id)
    supplier = procurement_store.get_project_supplier(supplier_id)
    if not project or not supplier or supplier['project_id'] != project_id:
        raise ValueError('采购项目或候选供应商不存在')
    filename = str(file_storage.filename or '')
    if not filename.lower().endswith('.xlsx'):
        raise ValueError('首版只支持标准 .xlsx 报价文件')
    if int(quote_round) < 1:
        raise ValueError('报价轮次必须大于等于 1')
    saved = procurement_file_service.save_upload(project, 'supplier_quote', file_storage)
    try:
        validate_office_archive(saved['absolute_path'])
        payload, errors, warnings = parse_standard_quote(
            saved['absolute_path'], project_id, supplier_id, int(quote_round)
        )
        payload['size_bytes'] = saved['size_bytes']
        return procurement_store.create_import_job({
            'project_id': project_id,
            'supplier_id': supplier_id,
            'quote_round': int(quote_round),
            'original_name': saved['original_name'],
            'relative_path': saved['relative_path'],
            'file_sha256': saved['sha256'],
            'parser_version': PARSER_VERSION,
            'payload': payload,
            'errors': errors,
            'warnings': warnings,
        })
    except Exception:
        try:
            Path(saved['absolute_path']).unlink(missing_ok=True)
        except OSError:
            get_logger().warning(
                '标准报价导入失败后无法删除上传文件',
                exc_info=True,
            )
        raise


def _validate_saved_pdf(path):
    path = Path(path)
    size = path.stat().st_size
    if size <= 0:
        raise ValueError('PDF 报价单不能为空')
    if size > MAX_QUOTE_PDF_BYTES:
        raise ValueError('PDF 报价单不能超过 25MB')
    with open(path, 'rb') as stream:
        header = stream.read(len(PDF_HEADER))
    if header != PDF_HEADER:
        raise ValueError('PDF 报价单文件内容无效，请上传真实 PDF 文件')


def save_quote_pdf_attachment(project_id, supplier_id, quote_round, file_storage):
    project = procurement_store.get_project(project_id)
    supplier = procurement_store.get_project_supplier(supplier_id)
    if not project or not supplier or supplier['project_id'] != project_id:
        raise ValueError('采购项目或候选供应商不存在')
    try:
        quote_round = int(quote_round)
    except (TypeError, ValueError) as exc:
        raise ValueError('报价轮次必须为整数') from exc
    if quote_round < 1:
        raise ValueError('报价轮次必须大于等于 1')
    filename = str(file_storage.filename or '').strip()
    if not filename:
        raise ValueError('请选择 PDF 报价单')
    if Path(filename).suffix.lower() != '.pdf':
        raise ValueError('PDF 报价单仅支持 .pdf 格式')

    saved = procurement_file_service.save_upload(project, 'supplier_quote_pdf', file_storage)
    try:
        _validate_saved_pdf(saved['absolute_path'])
        display_name = f"第{quote_round}轮_{supplier['supplier_name']}_{saved['original_name']}"
        return procurement_store.register_project_file(
            project_id, 'supplier_quote_pdf', saved['relative_path'], display_name,
            saved['sha256'], saved['size_bytes'],
        )
    except Exception:
        try:
            Path(saved['absolute_path']).unlink(missing_ok=True)
        except OSError:
            get_logger().warning(
                'PDF 报价保存失败后无法删除上传文件',
                exc_info=True,
            )
        raise


def confirm_import(job_id):
    return procurement_store.confirm_import_job(job_id)


def _validated_date(value, label):
    text = str(value or '').strip()
    if not text:
        return ''
    try:
        date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f'{label}格式无效') from exc
    return text


def _limited_form_text(form, key, label, max_length=2000):
    value = str(form.get(key) or '').strip()
    if len(value) > max_length:
        raise ValueError(f'{label}不能超过 {max_length} 个字符')
    return value


def _editable_tax_rate(form):
    raw = str(form.get('tax_rate') or '').strip()
    if not raw:
        return None
    try:
        value = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError('税率格式无效') from exc
    if not value.is_finite() or value < 0 or value > 100:
        raise ValueError('税率必须是 0 到 100 之间的有限数值')
    basis_points = value * 100
    if basis_points != basis_points.to_integral_value():
        raise ValueError('税率最多保留两位小数')
    return int(basis_points)


def update_confirmed_quote(quote_id, form):
    quote = procurement_store.get_quote(quote_id)
    if not quote:
        raise ValueError('供应商报价不存在')
    if quote['status'] != 'confirmed':
        raise ValueError('只有已确认的报价可以编辑')
    if quote.get('is_locked'):
        raise ValueError('该报价已用于成交建议，不能编辑')

    quote_date = _validated_date(form.get('quote_date'), '报价日期')
    quote_valid_until = _validated_date(form.get('quote_valid_until'), '报价有效期')
    if quote_date and quote_valid_until and quote_valid_until < quote_date:
        raise ValueError('报价有效期不能早于报价日期')
    price_basis = str(form.get('price_basis') or '').strip()
    if price_basis not in {'tax_inclusive', 'tax_exclusive'}:
        raise ValueError('价格口径无效')
    header = {
        'quote_date': quote_date,
        'quote_valid_until': quote_valid_until,
        'tax_rate_bps': _editable_tax_rate(form),
        'price_basis': price_basis,
        'delivery_period': _limited_form_text(form, 'delivery_period', '整体交付周期'),
        'payment_terms': _limited_form_text(form, 'payment_terms', '付款条件'),
        'warranty_period': _limited_form_text(form, 'warranty_period', '质保期'),
        'package_transport': _limited_form_text(form, 'package_transport', '包装运输'),
        'technical_deviation': _limited_form_text(
            form, 'technical_deviation', '整体技术偏离'
        ),
        'commercial_deviation': _limited_form_text(
            form, 'commercial_deviation', '整体商务偏离'
        ),
    }

    updated_items = []
    for item in procurement_store.get_quote_items(quote_id):
        item_id = item['id']
        raw_price = str(form.get(f'unit_price_{item_id}') or '').replace(',', '').strip()
        try:
            unit_price = Decimal(raw_price)
        except InvalidOperation as exc:
            raise ValueError(f'第 {item["line_no"]} 行单价格式无效') from exc
        if not unit_price.is_finite() or unit_price < 0:
            raise ValueError(f'第 {item["line_no"]} 行单价必须是非负有限数值')
        try:
            quantity = Decimal(item['quantity_text'])
        except InvalidOperation as exc:
            raise ValueError(f'第 {item["line_no"]} 行数量格式无效') from exc
        if not quantity.is_finite() or quantity <= 0:
            raise ValueError(f'第 {item["line_no"]} 行数量必须是正有限数值')
        unit_price_minor = to_minor(unit_price, allow_none=False)
        amount_minor = to_minor(quantity * unit_price, allow_none=False)
        updated_items.append({
            'id': item_id,
            'unit_price_minor': unit_price_minor,
            'amount_minor': amount_minor,
            'delivery_period': _limited_form_text(
                form, f'delivery_period_{item_id}', f'第 {item["line_no"]} 行交期', 1000
            ),
            'technical_deviation': _limited_form_text(
                form, f'technical_deviation_{item_id}',
                f'第 {item["line_no"]} 行技术偏离', 1000,
            ),
            'commercial_deviation': _limited_form_text(
                form, f'commercial_deviation_{item_id}',
                f'第 {item["line_no"]} 行商务偏离', 1000,
            ),
            'remark': _limited_form_text(
                form, f'remark_{item_id}', f'第 {item["line_no"]} 行备注', 1000
            ),
        })
    if not updated_items:
        raise ValueError('报价明细为空，不能保存')
    if sum(item['amount_minor'] for item in updated_items) > SQLITE_MAX_INTEGER:
        raise ValueError('报价总额超出可存储范围')
    return procurement_store.update_quote(quote_id, header, updated_items)


def delete_confirmed_quote(quote_id):
    result = procurement_store.delete_quote(quote_id)
    relative_path = result.get('relative_path') or ''
    if relative_path:
        try:
            procurement_file_service.absolute_path(relative_path).unlink(missing_ok=True)
        except (OSError, ValueError):
            get_logger().warning(
                '删除供应商报价原文件失败: %s', relative_path, exc_info=True
            )
    return result['project_id']
