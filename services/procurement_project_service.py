"""Procurement project application service."""

from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import ledger_store
import procurement_store
from services import procurement_file_service
from utils import helpers
from utils.constants import (
    PROCUREMENT_METHOD_LABELS,
    PROCUREMENT_STAGE_LABELS,
    PROCUREMENT_STAGE_ORDER,
    PROCUREMENT_STATUS_LABELS,
)


STATUS_TRANSITIONS = {
    'draft': {'documents_ready', 'inquiry_sent', 'quotes_received', 'archived'},
    'documents_ready': {'draft', 'inquiry_sent', 'quotes_received', 'archived'},
    'inquiry_sent': {'documents_ready', 'quotes_received', 'archived'},
    'quotes_received': {'inquiry_sent', 'clarifying', 'negotiating', 'award_draft', 'award_confirmed', 'archived'},
    'clarifying': {'quotes_received', 'negotiating', 'award_draft', 'award_confirmed', 'archived'},
    'negotiating': {'quotes_received', 'clarifying', 'award_draft', 'award_confirmed', 'archived'},
    'award_draft': {'quotes_received', 'award_confirmed', 'archived'},
    'award_confirmed': {'award_draft', 'contract_draft', 'archived'},
    'contract_draft': {'award_confirmed', 'contract_created', 'archived'},
    'contract_created': {'archived'},
    'archived': {'draft', 'contract_created'},
}

STAGE_STATUS_MAP = {
    'project': 'draft',
    'items': 'draft',
    'suppliers': 'draft',
    'quotes': 'quotes_received',
    'comparison': 'clarifying',
    'negotiation': 'negotiating',
    'award': 'award_draft',
    'contract': 'contract_draft',
    'archive': 'archived',
}

STAGE_ACTIONS = {
    'project': {'endpoint': 'procurement_project_edit', 'label': '编辑项目'},
    'items': {'endpoint': 'procurement_project_detail', 'anchor': 'items', 'label': '录入明细'},
    'suppliers': {'endpoint': 'procurement_project_detail', 'anchor': 'suppliers', 'label': '维护供应商'},
    'quotes': {'endpoint': 'procurement_quote_import', 'label': '导入报价'},
    'comparison': {'endpoint': 'procurement_comparison', 'label': '进入比价'},
    'negotiation': {'endpoint': 'procurement_negotiation', 'label': '进入谈判'},
    'award': {'endpoint': 'procurement_award', 'label': '成交建议'},
    'contract': {'endpoint': 'procurement_direct_contract', 'label': '直接生成合同'},
    'archive': {'endpoint': 'procurement_project_archive', 'label': '生成归档包'},
}


def money_to_minor(value, label='金额', allow_empty=True):
    raw = str(value or '').replace(',', '').strip()
    if not raw and allow_empty:
        return None
    try:
        amount = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f'{label}格式无效') from exc
    if amount < 0:
        raise ValueError(f'{label}不能为负数')
    return int((amount * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def minor_to_money(value):
    if value is None:
        return ''
    return f'{Decimal(int(value)) / 100:.2f}'


def _next_project_no():
    prefix = 'CG-' + date.today().strftime('%Y%m%d') + '-'
    with ledger_store.get_conn() as conn:
        conn.execute('BEGIN IMMEDIATE')
        rows = conn.execute(
            'SELECT project_no FROM procurement_projects WHERE project_no LIKE ? ORDER BY project_no DESC LIMIT 1',
            (prefix + '%',),
        ).fetchone()
    if not rows:
        return prefix + '0001'
    try:
        sequence = int(rows[0].rsplit('-', 1)[1]) + 1
    except (ValueError, IndexError):
        sequence = 1
    return prefix + f'{sequence:04d}'


def create_project(form):
    import sqlite3
    name = str(form.get('project_name') or '').strip()
    if not name:
        raise ValueError('项目名称不能为空')
    project_no = str(form.get('project_no') or '').strip() or _next_project_no()
    data = {
        'project_no': project_no,
        'project_name': name,
        'purchase_method': str(form.get('purchase_method') or 'competitive_negotiation').strip(),
        'demand_department': str(form.get('demand_department') or '').strip(),
        'owner': str(form.get('owner') or '').strip(),
        'budget_minor': money_to_minor(form.get('budget_amount'), '预算金额'),
        'target_price_minor': money_to_minor(form.get('target_price'), '目标价格'),
        'currency': 'CNY',
        'delivery_place': str(form.get('delivery_place') or '').strip(),
        'delivery_requirement': str(form.get('delivery_requirement') or '').strip(),
        'payment_requirement': str(form.get('payment_requirement') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    }
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if attempt > 0:
                project_no = _next_project_no()
                data['project_no'] = project_no
            project_id = procurement_store.create_project(data)
            break
        except sqlite3.IntegrityError:
            if attempt == max_retries - 1:
                raise ValueError('创建项目失败，请稍后重试')
    project = procurement_store.get_project(project_id)
    procurement_file_service.ensure_project_folders(project)
    return project_id


def update_project(project_id, form):
    project = procurement_store.get_project(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    name = str(form.get('project_name') or '').strip()
    if not name:
        raise ValueError('项目名称不能为空')
    return procurement_store.update_project(project_id, {
        'project_name': name,
        'purchase_method': str(form.get('purchase_method') or project['purchase_method']).strip(),
        'demand_department': str(form.get('demand_department') or '').strip(),
        'owner': str(form.get('owner') or '').strip(),
        'budget_minor': money_to_minor(form.get('budget_amount'), '预算金额'),
        'target_price_minor': money_to_minor(form.get('target_price'), '目标价格'),
        'delivery_place': str(form.get('delivery_place') or '').strip(),
        'delivery_requirement': str(form.get('delivery_requirement') or '').strip(),
        'payment_requirement': str(form.get('payment_requirement') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    })


def transition(project_id, new_status, note=''):
    project = procurement_store.get_project(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    if new_status == project['status']:
        return
    if new_status not in STATUS_TRANSITIONS.get(project['status'], set()):
        current_label = PROCUREMENT_STATUS_LABELS.get(project['status'], project['status'])
        new_label = PROCUREMENT_STATUS_LABELS.get(new_status, new_status)
        raise ValueError(f'项目不能从“{current_label}”直接变更为“{new_label}”')
    procurement_store.transition_project_status(project_id, new_status, note=note)


def _stage_completion(project):
    return {
        'project': True,
        'items': len(project.get('items') or []) > 0,
        'suppliers': len(project.get('suppliers') or []) > 0,
        'quotes': len(project.get('quotes') or []) > 0,
        'comparison': bool(project.get('comparison') or project.get('clarifications')),
        'negotiation': len(project.get('negotiation_rounds') or []) > 0,
        'award': bool(project.get('award')),
        'contract': len(project.get('contract_links') or []) > 0,
        'archive': project.get('status') == 'archived',
    }


def _workflow_skips(project_id):
    skipped = {}
    for event in procurement_store.list_project_audit_events(project_id, actions=['workflow_jump']):
        after = event.get('after') or {}
        note = event.get('note') or ''
        for stage in after.get('skipped_stages') or []:
            skipped.setdefault(stage, {
                'note': note,
                'target_stage': after.get('target_stage') or '',
                'created_at': event.get('created_at') or '',
            })
    return skipped


def _missing_before(project, target_stage):
    completion = _stage_completion(project)
    missing = []
    for stage in PROCUREMENT_STAGE_ORDER:
        if stage == target_stage:
            break
        if stage == 'project':
            continue
        if not completion.get(stage):
            missing.append(stage)
    return missing


def build_workflow_view(project_id):
    project = project_detail(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    completion = _stage_completion(project)
    skipped = _workflow_skips(project_id)
    recommended_key = None
    for stage in PROCUREMENT_STAGE_ORDER:
        if not completion.get(stage) and stage not in skipped:
            recommended_key = stage
            break
    if recommended_key is None:
        recommended_key = 'archive'

    stages = []
    for stage in PROCUREMENT_STAGE_ORDER:
        done = completion.get(stage, False)
        skipped_info = skipped.get(stage)
        missing_before = _missing_before(project, stage)
        if done:
            status = 'done'
        elif skipped_info:
            status = 'skipped'
        elif stage == recommended_key:
            status = 'active'
        elif missing_before:
            status = 'available'
        else:
            status = 'available'
        stages.append({
            'key': stage,
            'label': PROCUREMENT_STAGE_LABELS.get(stage, stage),
            'status': status,
            'done': done,
            'skipped': bool(skipped_info),
            'skip_note': (skipped_info or {}).get('note', ''),
            'missing_before': missing_before,
            'missing_labels': [PROCUREMENT_STAGE_LABELS.get(item, item) for item in missing_before],
            'action': STAGE_ACTIONS.get(stage, {}),
        })
    return {
        'status_label': PROCUREMENT_STATUS_LABELS.get(project.get('status'), project.get('status')),
        'method_label': PROCUREMENT_METHOD_LABELS.get(
            project.get('purchase_method'), project.get('purchase_method')
        ),
        'recommended_stage': recommended_key,
        'recommended_label': PROCUREMENT_STAGE_LABELS.get(recommended_key, recommended_key),
        'stages': stages,
    }


def jump_to_stage(project_id, target_stage, note=''):
    if target_stage not in PROCUREMENT_STAGE_ORDER:
        raise ValueError('采购环节无效')
    project = project_detail(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    if target_stage == 'negotiation' and not project.get('suppliers'):
        raise ValueError('进入谈判前至少需要添加一个候选供应商')
    note = str(note or '').strip()
    missing = _missing_before(project, target_stage)
    if missing and not note:
        raise ValueError('跳过前置环节时需要填写原因')
    before_status = project.get('status') or ''
    next_status = STAGE_STATUS_MAP.get(target_stage, before_status)
    procurement_store.transition_project_status(project_id, next_status, note=note)
    procurement_store.record_workflow_jump(
        project_id, target_stage, missing, note=note,
        before_status=before_status, after_status=next_status,
    )
    return target_stage


def add_item(project_id, form):
    if not procurement_store.get_project(project_id):
        raise ValueError('采购项目不存在')
    item_name = str(form.get('item_name') or '').strip()
    unit = str(form.get('unit') or '').strip()
    if not item_name:
        raise ValueError('物资名称不能为空')
    if not unit:
        raise ValueError('单位不能为空')
    try:
        quantity = Decimal(str(form.get('quantity') or '').strip())
    except InvalidOperation as exc:
        raise ValueError('数量格式无效') from exc
    if quantity <= 0:
        raise ValueError('数量必须大于 0')
    normalized = format(quantity.normalize(), 'f')
    return procurement_store.add_project_item(project_id, {
        'item_name': item_name,
        'spec_model': str(form.get('spec_model') or '').strip(),
        'drawing_no': str(form.get('drawing_no') or '').strip(),
        'quantity_text': normalized,
        'unit': unit,
        'required_delivery_date': str(form.get('required_delivery_date') or '').strip(),
        'technical_requirement': str(form.get('technical_requirement') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    })


def update_item(project_id, item_id, form):
    item_name = str(form.get('item_name') or '').strip()
    unit = str(form.get('unit') or '').strip()
    if not item_name or not unit:
        raise ValueError('物资名称和单位不能为空')
    try:
        quantity = Decimal(str(form.get('quantity') or '').strip())
    except InvalidOperation as exc:
        raise ValueError('数量格式无效') from exc
    if quantity <= 0:
        raise ValueError('数量必须大于 0')
    procurement_store.update_project_item(project_id, item_id, {
        'item_name': item_name,
        'spec_model': str(form.get('spec_model') or '').strip(),
        'drawing_no': str(form.get('drawing_no') or '').strip(),
        'quantity_text': format(quantity.normalize(), 'f'),
        'unit': unit,
        'required_delivery_date': str(form.get('required_delivery_date') or '').strip(),
        'technical_requirement': str(form.get('technical_requirement') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    })


def add_items_from_rows(project_id, rows):
    parsed = []
    errors = []
    for row_number, row in enumerate(rows, start=1):
        values = list(row) + [''] * 8
        if not any(str(value or '').strip() for value in values):
            continue
        item_name = str(values[0] or '').strip()
        unit = str(values[4] or '').strip()
        try:
            quantity = Decimal(str(values[3] or '').strip())
        except InvalidOperation:
            errors.append(f'第 {row_number} 行数量格式无效')
            continue
        if not item_name or not unit or quantity <= 0:
            errors.append(f'第 {row_number} 行物资名称、正数数量和单位为必填项')
            continue
        parsed.append({
            'item_name': item_name, 'spec_model': str(values[1] or '').strip(),
            'drawing_no': str(values[2] or '').strip(),
            'quantity_text': format(quantity.normalize(), 'f'), 'unit': unit,
            'required_delivery_date': str(values[5] or '').strip(),
            'technical_requirement': str(values[6] or '').strip(),
            'remark': str(values[7] or '').strip(),
        })
    if errors:
        raise ValueError('；'.join(errors[:20]))
    if not parsed:
        raise ValueError('没有可导入的采购明细')
    return procurement_store.add_project_items_bulk(project_id, parsed)


def add_items_from_paste(project_id, text):
    rows = [line.split('\t') for line in str(text or '').splitlines()]
    return add_items_from_rows(project_id, rows)


def add_items_from_excel(project_id, file_storage):
    from openpyxl import load_workbook
    workbook = load_workbook(file_storage.stream, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if rows and str(rows[0][0] or '').strip() in {'物资名称', '产品名称', '标的名称'}:
        rows = rows[1:]
    return add_items_from_rows(project_id, rows)


def add_supplier(project_id, form):
    if not procurement_store.get_project(project_id):
        raise ValueError('采购项目不存在')
    supplier_name = str(form.get('supplier_name') or '').strip()
    if not supplier_name:
        raise ValueError('供应商名称不能为空')
    return procurement_store.add_project_supplier(project_id, {
        'supplier_name': supplier_name,
        'contact_person': str(form.get('contact_person') or '').strip(),
        'contact_phone': str(form.get('contact_phone') or '').strip(),
        'email': str(form.get('email') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    })


def update_supplier(project_id, supplier_id, form):
    supplier_name = str(form.get('supplier_name') or '').strip()
    if not supplier_name:
        raise ValueError('供应商名称不能为空')
    procurement_store.update_project_supplier(project_id, supplier_id, {
        'supplier_name': supplier_name,
        'contact_person': str(form.get('contact_person') or '').strip(),
        'contact_phone': str(form.get('contact_phone') or '').strip(),
        'email': str(form.get('email') or '').strip(),
        'remark': str(form.get('remark') or '').strip(),
    })


def prepare_direct_contract_session(project_id, template_filename):
    project = project_detail(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    path = helpers.safe_template_path(template_filename)
    primary_supplier = (project.get('suppliers') or [{}])[0]
    amount_minor = project.get('target_price_minor') or project.get('budget_minor') or 0
    payload = {
        'schema_version': 'direct-1.0',
        'project_id': project_id,
        'project_no': project['project_no'],
        'project_name': project['project_name'],
        'award_recommendation_id': None,
        'supplier': {
            'id': primary_supplier.get('id'),
            'name': primary_supplier.get('supplier_name') or '',
        },
        'suppliers': [row.get('supplier_name') for row in project.get('suppliers', []) if row.get('supplier_name')],
        'currency': project.get('currency') or 'CNY',
        'amount_minor': amount_minor,
        'items': [
            {
                'project_item_id': item['id'],
                'item_name': item['item_name'],
                'spec_model': item.get('spec_model') or '',
                'quantity': item.get('quantity_text') or '',
                'unit': item.get('unit') or '',
                'unit_price': '',
                'amount': '',
                'supplier_name': primary_supplier.get('supplier_name') or '',
            }
            for item in project.get('items', [])
        ],
        'delivery_place': project.get('delivery_place') or '',
        'delivery_terms': project.get('delivery_requirement') or '',
        'payment_terms': project.get('payment_requirement') or '',
        'warranty_period': '',
        'technical_notes': '',
        'commercial_notes': '',
        'contract_notice': project.get('remark') or '',
        'owner': project.get('owner') or '',
        'demand_department': project.get('demand_department') or '',
        'source_quote_ids': [],
    }
    from services import award_service
    tpl, fields = award_service.prepare_template_fields(path, payload)
    return {
        'template_name': tpl.name,
        'template_path': path,
        'template_filename': template_filename,
        'stored_name': tpl.data.get('source_docx', ''),
        'fields': fields,
        'initial_values': payload,
        'source_project_id': project_id,
        'source_type': 'direct_contract',
        'procurement_project_id': project_id,
        'project_name': project['project_name'],
        'step': 'editor',
    }


def project_detail(project_id):
    project = procurement_store.get_project(project_id)
    if not project:
        return None
    project['budget_amount'] = minor_to_money(project.get('budget_minor'))
    project['target_price'] = minor_to_money(project.get('target_price_minor'))
    project['items'] = procurement_store.list_project_items(project_id)
    project['suppliers'] = procurement_store.list_project_suppliers(project_id)
    project['quotes'] = procurement_store.list_quotes(project_id)
    project['clarifications'] = procurement_store.list_clarifications(project_id)
    project['comparison'] = procurement_store.get_latest_comparison(project_id)
    project['negotiation_rounds'] = procurement_store.list_negotiation_rounds(project_id)
    project['award'] = procurement_store.get_latest_award(project_id)
    project['contract_links'] = procurement_store.get_project_contract_links(project_id)
    project['files'] = procurement_store.list_project_files(project_id)
    return project
