"""Quotation comparison, anomaly rules, and clarification drafts."""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import procurement_store
from services import procurement_file_service
from utils.money import from_minor


DEFAULT_THRESHOLD_PERCENT = Decimal('20')


def _money(value):
    return from_minor(value or 0)


def _latest_quote_data(project_id):
    quotes = procurement_store.get_latest_quotes(project_id)
    for quote in quotes:
        quote['items'] = procurement_store.get_quote_items(quote['id'])
        quote['items_by_project'] = {item['project_item_id']: item for item in quote['items']}
    return quotes


def run_comparison(project_id, threshold_percent=None):
    project = procurement_store.get_project(project_id)
    items = procurement_store.list_project_items(project_id)
    quotes = _latest_quote_data(project_id)
    if not project:
        raise ValueError('采购项目不存在')
    if not items:
        raise ValueError('采购项目没有明细')
    if not quotes:
        raise ValueError('请先导入并确认至少一份报价')
    config = procurement_store.get_rule_config(project_id)
    threshold_percent = (threshold_percent if threshold_percent is not None
                         else config['price_threshold_percent'])
    threshold = Decimal(str(threshold_percent)) / 100
    min_valid_suppliers = max(2, int(config.get('min_valid_suppliers') or 2))
    results = []
    price_bases = {quote.get('price_basis') for quote in quotes}
    comparable_price_basis = (len(price_bases) == 1 or not config.get('require_same_price_basis'))

    for project_item in items:
        present = []
        for quote in quotes:
            quote_item = quote['items_by_project'].get(project_item['id'])
            if not quote_item:
                results.append({
                    'project_item_id': project_item['id'], 'supplier_id': quote['supplier_id'],
                    'quote_id': quote['id'], 'result_type': 'missing_item', 'severity': 'high',
                    'description': f"{quote['supplier_name']} 未报价：{project_item['item_name']}",
                    'suggestion': '请供应商补充该项报价',
                })
                continue
            present.append((quote, quote_item))
            if quote_item.get('technical_deviation'):
                results.append({
                    'project_item_id': project_item['id'], 'supplier_id': quote['supplier_id'],
                    'quote_id': quote['id'], 'result_type': 'technical_deviation', 'severity': 'medium',
                    'description': f"{quote['supplier_name']} 技术偏离：{quote_item['technical_deviation']}",
                    'suggestion': '确认偏离是否可接受',
                })
            if quote_item.get('commercial_deviation'):
                results.append({
                    'project_item_id': project_item['id'], 'supplier_id': quote['supplier_id'],
                    'quote_id': quote['id'], 'result_type': 'commercial_deviation', 'severity': 'medium',
                    'description': f"{quote['supplier_name']} 商务偏离：{quote_item['commercial_deviation']}",
                    'suggestion': '确认商务偏离处理方式',
                })
        positive = [Decimal(item['unit_price_minor']) for _, item in present if item['unit_price_minor'] > 0]
        if comparable_price_basis and len(positive) >= min_valid_suppliers:
            average = sum(positive) / len(positive)
            for quote, quote_item in present:
                price = Decimal(quote_item['unit_price_minor'])
                if price <= 0:
                    continue
                deviation = (price - average) / average
                if deviation > threshold:
                    result_type, severity, suggestion = 'high_price', 'medium', '请确认高价原因或进一步议价'
                elif deviation < -threshold:
                    result_type, severity, suggestion = 'low_price', 'medium', '请确认报价完整性和履约可行性'
                else:
                    continue
                results.append({
                    'project_item_id': project_item['id'], 'supplier_id': quote['supplier_id'],
                    'quote_id': quote['id'], 'result_type': result_type, 'severity': severity,
                    'description': (
                        f"{quote['supplier_name']} {project_item['item_name']}单价 {_money(price)}，"
                        f"偏离均值 {_money(average)} 的 {deviation * 100:.2f}%"
                    ),
                    'suggestion': suggestion,
                    'metric': {'unit_price_minor': int(price), 'average_minor': int(average),
                               'deviation_percent': float(deviation * 100)},
                })

    tax_rates = {quote.get('tax_rate_bps') for quote in quotes if quote.get('tax_rate_bps') is not None}
    for quote in quotes:
        if len(tax_rates) > 1:
            results.append({
                'supplier_id': quote['supplier_id'], 'quote_id': quote['id'],
                'result_type': 'tax_rate_mismatch', 'severity': 'medium',
                'description': f"{quote['supplier_name']} 税率与其他供应商不一致",
                'suggestion': '确认统一税率或换算口径',
            })
        if len(price_bases) > 1:
            results.append({
                'supplier_id': quote['supplier_id'], 'quote_id': quote['id'],
                'result_type': 'price_basis_mismatch', 'severity': 'high',
                'description': f"{quote['supplier_name']} 含税/不含税口径不一致",
                'suggestion': '统一价格口径后重新比价',
            })
        if project.get('delivery_requirement') and quote.get('delivery_period') and (
            project['delivery_requirement'].strip() != quote['delivery_period'].strip()
        ):
            results.append({
                'supplier_id': quote['supplier_id'], 'quote_id': quote['id'],
                'result_type': 'delivery_deviation', 'severity': 'medium',
                'description': f"{quote['supplier_name']} 交付周期与项目要求不同：{quote['delivery_period']}",
                'suggestion': '确认供应商能否满足项目交付要求',
            })
        if project.get('payment_requirement') and quote.get('payment_terms') and (
            project['payment_requirement'].strip() != quote['payment_terms'].strip()
        ):
            results.append({
                'supplier_id': quote['supplier_id'], 'quote_id': quote['id'],
                'result_type': 'payment_deviation', 'severity': 'medium',
                'description': f"{quote['supplier_name']} 付款条件与项目要求不同：{quote['payment_terms']}",
                'suggestion': '澄清并确认最终付款条件',
            })
    run_id = procurement_store.create_comparison_run(
        project_id, [quote['id'] for quote in quotes],
        {'threshold_percent': float(Decimal(str(threshold_percent))),
         'min_valid_suppliers': min_valid_suppliers,
         'require_same_price_basis': bool(config.get('require_same_price_basis'))}, results,
    )
    return run_id


def comparison_view(project_id):
    project = procurement_store.get_project(project_id)
    items = procurement_store.list_project_items(project_id)
    quotes = _latest_quote_data(project_id)
    rows = []
    for item in items:
        row = {'item': item, 'quotes': {}}
        for quote in quotes:
            row['quotes'][quote['id']] = quote['items_by_project'].get(item['id'])
        rows.append(row)
    return {
        'project': project,
        'quotes': quotes,
        'rows': rows,
        'comparison': procurement_store.get_latest_comparison(project_id),
        'rule_config': procurement_store.get_rule_config(project_id),
    }


def generate_clarifications(project_id):
    comparison = procurement_store.get_latest_comparison(project_id)
    if not comparison:
        raise ValueError('请先执行横向比价')
    templates = {
        'missing_item': '请补充“{item}”的报价，并确认是否包含在总报价中。',
        'high_price': '“{item}”报价明显高于其他供应商，请说明价格构成及进一步优惠空间。',
        'low_price': '“{item}”报价明显偏低，请确认报价完整性、技术响应及履约可行性。',
        'technical_deviation': '请对“{item}”的技术偏离逐项说明，并确认能否按我方要求执行。',
        'commercial_deviation': '请对“{item}”的商务偏离逐项说明，并提出可接受的处理方案。',
        'delivery_deviation': '请确认最终交付周期以及能否满足项目要求。',
        'payment_deviation': '请确认最终付款条件以及能否接受我方付款要求。',
        'tax_rate_mismatch': '请确认本次报价税率及含税金额计算口径。',
        'price_basis_mismatch': '请确认本次报价为含税或不含税，并按统一口径重新报价。',
    }
    questions = []
    for result in comparison['results']:
        if result['status'] == 'ignored' or result['result_type'] not in templates:
            continue
        questions.append({
            'supplier_id': result.get('supplier_id'),
            'project_item_id': result.get('project_item_id'),
            'question_type': result['result_type'],
            'question_text': templates[result['result_type']].format(item=result.get('item_name') or '相关项目'),
            'source_result_id': result['id'],
        })
    return procurement_store.create_clarifications_from_results(project_id, questions)


def export_comparison_excel(project_id):
    view = comparison_view(project_id)
    if not view['quotes']:
        raise ValueError('没有可导出的报价')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '横向比价'
    headers = ['行号', '物资名称', '规格型号', '数量', '单位']
    for quote in view['quotes']:
        headers.extend([f"{quote['supplier_name']}-单价", f"{quote['supplier_name']}-金额"])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
    for row in view['rows']:
        item = row['item']
        values = [item['line_no'], item['item_name'], item['spec_model'], item['quantity_text'], item['unit']]
        for quote in view['quotes']:
            quote_item = row['quotes'].get(quote['id'])
            values.extend([
                Decimal(quote_item['unit_price_minor']) / 100 if quote_item else '',
                Decimal(quote_item['amount_minor']) / 100 if quote_item else '',
            ])
        sheet.append(values)
    anomaly = workbook.create_sheet('异常清单')
    anomaly.append(['类型', '供应商', '物资', '严重度', '说明', '建议'])
    if view['comparison']:
        for result in view['comparison']['results']:
            anomaly.append([
                result['result_type'], result.get('supplier_name') or '', result.get('item_name') or '',
                result['severity'], result['description'], result.get('suggestion') or '',
            ])
    project = view['project']
    path = procurement_file_service.target_path(
        project, 'comparison', f"{project['project_no']}_横向比价.xlsx"
    )
    workbook.save(path)
    procurement_store.register_project_file(
        project_id, 'comparison', procurement_file_service.relative_path(path), path.name,
        procurement_file_service.sha256_file(path), path.stat().st_size,
    )
    return str(path)
