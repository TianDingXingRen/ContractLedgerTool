"""XLSX 导出模块 — 基于 openpyxl 生成付款计划和合同台账 Excel 文件"""

import math
import os
import re
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from utils.security import safe_spreadsheet_value
from xlsx_export.columns import (
    CONTRACT_COLUMNS,
    HANDOVER_OVERVIEW_COLUMNS,
    HANDOVER_TABLE_COLUMNS,
    MONEY_FORMAT,
    MONTHLY_BASE_COLUMNS,
    MONTHLY_SUMMARY_COLUMNS,
    PAYMENT_PLAN_COLUMNS,
    column_headers,
    money_column_numbers,
    monthly_detail_columns,
    monthly_summary_columns,
)
from xlsx_export.styles import WorkbookStyleFactory


_STYLES = WorkbookStyleFactory()


def _apply_header_style(ws, row, col_count):
    _STYLES.apply_header_row(ws, row, col_count)


def _num(value):
    try:
        v = float(value or 0)
        return v if math.isfinite(v) else 0
    except (TypeError, ValueError):
        return 0


def export_payment_plans(path, rows, title='下月付款计划'):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    columns = PAYMENT_PLAN_COLUMNS
    headers = column_headers(columns)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(
        row=1, column=1, value=safe_spreadsheet_value(title)
    )
    title_cell.font = _STYLES.title_font
    title_cell.alignment = _STYLES.title_alignment

    # 表头（第3行）
    for ci, header in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=header)
    _apply_header_style(ws, 3, len(headers))

    # 数据行（从第4行开始）
    total_due = total_paid = 0
    for idx, row in enumerate(rows, 1):
        ri = idx + 3  # Excel 行号
        due = _num(row.get('due_amount'))
        paid = _num(row.get('paid_amount'))
        total_due += due
        total_paid += paid
        values = [
            idx,
            row.get('project_name') or '',
            row.get('subsystem_name') or row.get('contract_subsystem_name') or '',
            (
                (
                    f"第 {row.get('serial_no')} 发（历史关联）"
                    if row.get('coverage_not_applicable')
                    else f"第 {row.get('serial_no')} 发"
                )
                if row.get('serial_no') is not None
                else '不适用' if row.get('coverage_not_applicable')
                else '待补发次'
            ),
            (f"第 {row.get('coverage_start')}–{row.get('coverage_end')} 发"
             if row.get('coverage_start') is not None and row.get('coverage_end') is not None
             else '不适用' if row.get('coverage_not_applicable') else '待补发次'),
            row.get('contract_no') or '',
            row.get('contract_title') or '',
            row.get('counterparty') or '',
            row.get('phase_name') or '',
            row.get('due_date') or '',
            due, paid, round(due - paid, 2),
            row.get('condition_text') or '',
            row.get('owner') or '',
            row.get('remark') or '',
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=safe_spreadsheet_value(val))
            cell.font = _STYLES.normal_font
            cell.border = _STYLES.thin_border
            if ci in (11, 12, 13):
                cell.number_format = MONEY_FORMAT

    # 合计行
    summary_row = len(rows) + 4
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=10)
    if rows:
        ws.cell(row=summary_row, column=1, value='合计').font = _STYLES.header_font
        ws.cell(row=summary_row, column=11, value=round(total_due, 2)).number_format = MONEY_FORMAT
        ws.cell(row=summary_row, column=12, value=round(total_paid, 2)).number_format = MONEY_FORMAT
        ws.cell(row=summary_row, column=13, value=round(total_due - total_paid, 2)).number_format = MONEY_FORMAT
    else:
        ws.cell(row=summary_row, column=1, value='（无数据）').font = _STYLES.header_font
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=ci)
        cell.border = _STYLES.thin_border

    _STYLES.apply_column_widths(ws, columns)

    try:
        wb.save(path)
    finally:
        wb.close()
    return path


def _minor_to_wan(value):
    if value is None:
        return None
    amount = Decimal(int(value)) / Decimal(1_000_000)
    return float(amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


def _safe_sheet_name(value, used):
    name = re.sub(r'[\[\]:*?/\\]', '_', str(value or '未归类项目')).strip() or '未归类项目'
    name = name[:31]
    candidate = name
    suffix = 2
    used_casefolded = {str(item).casefold() for item in used}
    while candidate.casefold() in used_casefolded:
        tail = f'-{suffix}'
        candidate = f'{name[:31-len(tail)]}{tail}'
        suffix += 1
    used.add(candidate)
    return candidate


def _wrapped_line_count(value, column_width):
    """Estimate wrapped lines for mixed Chinese and Latin workbook text."""
    text = str(value or '')
    if not text:
        return 1
    capacity = max(4, int(float(column_width or 10) * 0.9))
    line_count = 0
    for segment in text.splitlines() or ['']:
        display_units = sum(2 if ord(char) > 127 else 1 for char in segment)
        line_count += max(1, math.ceil(display_units / capacity))
    return line_count


def _monthly_detail_row_height(values, columns):
    max_lines = max(
        _wrapped_line_count(value, definition.width)
        for value, definition in zip(values, columns)
    )
    return max(33, min(90, max_lines * 15 + 6))


def _monthly_summary_row_height(project_name, subsystem_name=''):
    lines = max(
        _wrapped_line_count(project_name, MONTHLY_SUMMARY_COLUMNS[0].width),
        _wrapped_line_count(subsystem_name, MONTHLY_SUMMARY_COLUMNS[1].width),
    )
    return max(20, min(80, lines * 20 + 6))


def _style_monthly_sheet(
    ws,
    last_col,
    data_start,
    reserved_data_end,
    planned_total_row,
):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = 'F4'
    ws.row_dimensions[1].height = 67
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 60
    for row_index in range(data_start, reserved_data_end + 1):
        if ws.row_dimensions[row_index].height is None:
            ws.row_dimensions[row_index].height = 33
    for row_index in range(planned_total_row - 2, planned_total_row + 1):
        ws.row_dimensions[row_index].height = 30
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.print_title_rows = '1:3'


def export_monthly_payment_plan_report(path, report):
    """Export the reference-style monthly workbook in ten-thousand yuan.

    ``report`` is built by ``ledger_store.build_monthly_payment_report`` and
    intentionally contains no production-notice data.
    """
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = '汇总'
    used_sheet_names = {'汇总'}
    month_text = str(report['report_month'])
    year, month = month_text.split('-')
    display_month = f'{int(year)}年{int(month)}月'
    node_count = max(1, int(report.get('node_count') or 1))
    project_refs = []

    for project in report.get('projects', []):
        sheet_name = _safe_sheet_name(
            f"{project['project_name']}-{project['subsystem_name']}",
            used_sheet_names,
        )
        ws = wb.create_sheet(sheet_name)
        sheet_name = ws.title
        used_sheet_names.add(sheet_name)
        columns = monthly_detail_columns(node_count)
        headers = column_headers(columns)
        last_col = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws.cell(
            1, 1,
            '填表注意事项：1、数据万元显示；2、数据右对齐，保留2位；'
            '3、每个火箭型号项目的各分系统分别生成明细页，并在汇总页汇总；'
            '4、如预计在20日前无法完成单据审批，则将付款计划做到次月；'
            '5、如可用银承支付，明确银承兑付时间；',
        )
        ws.cell(1, 1).font = _STYLES.monthly_note_font
        ws.cell(1, 1).alignment = _STYLES.left_wrap_alignment
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws.cell(2, 1, f'{display_month}付款明细表    单位：万元')
        ws.cell(2, 1).font = _STYLES.monthly_title_font
        ws.cell(2, 1).alignment = _STYLES.center_alignment
        for column, header in enumerate(headers, 1):
            ws.cell(3, column, safe_spreadsheet_value(header))
            ws.cell(3, column).font = _STYLES.monthly_header_font
            ws.cell(3, column).fill = _STYLES.monthly_header_fill
            ws.cell(3, column).alignment = _STYLES.center_wrap_alignment
            ws.cell(3, column).border = _STYLES.thin_border

        current_payment_col = len(MONTHLY_BASE_COLUMNS) + node_count * 2 + 1
        overdue_col = current_payment_col + 2
        bank_col = current_payment_col + 3
        helper_current_col = last_col + 1
        helper_previous_col = last_col + 2
        helper_current_letter = get_column_letter(helper_current_col)
        helper_previous_letter = get_column_letter(helper_previous_col)
        current_payment_letter = get_column_letter(current_payment_col)
        bank_letter = get_column_letter(bank_col)
        ws.column_dimensions[helper_current_letter].hidden = True
        ws.column_dimensions[helper_previous_letter].hidden = True

        data_start = 4
        data_end = data_start + len(project['rows']) - 1
        reserved_data_end = max(10, data_end)
        for row_index in range(data_start, reserved_data_end + 1):
            for column in range(1, last_col + 1):
                cell = ws.cell(row_index, column)
                cell.font = _STYLES.monthly_body_font
                cell.border = _STYLES.thin_border
                cell.alignment = _STYLES.center_wrap_alignment
            ws.cell(row_index, current_payment_col, 0)
            ws.cell(row_index, current_payment_col).number_format = MONEY_FORMAT

        for row_index, row in enumerate(project['rows'], data_start):
            values = [
                (
                    (
                        f"第 {row['serial_no']} 发（历史关联；合同发次不适用）"
                        if row.get('coverage_not_applicable')
                        else row['serial_no']
                    )
                    if row.get('serial_no') is not None
                    else '不适用'
                    if row.get('coverage_not_applicable')
                    else '待补发次'
                ),
                row['contract_no'], row['contract_title'],
                row['party_a'], row['party_b'],
                _minor_to_wan(row['serial_amount_minor']),
            ]
            for node_index in range(node_count):
                node = row['nodes'][node_index] if node_index < len(row['nodes']) else None
                values.extend([
                    _minor_to_wan(node['amount_minor']) if node else None,
                    node['condition'] if node else '',
                ])
            values.extend([
                None,
                row['payment_condition'],
                row['overdue_text'],
                row['bank_acceptance'],
                row['prior_unpaid_reason'],
            ])
            for column, value in enumerate(values, 1):
                cell = ws.cell(row_index, column, safe_spreadsheet_value(value))
                cell.font = _STYLES.monthly_body_font
                cell.border = _STYLES.thin_border
                cell.alignment = (
                    _STYLES.right_wrap_alignment
                    if (
                        column == 6
                        or (
                            7 <= column < current_payment_col
                            and column % 2 == 1
                        )
                        or column == current_payment_col
                    )
                    else _STYLES.center_wrap_alignment
                )
            ws.row_dimensions[row_index].height = _monthly_detail_row_height(
                values,
                columns,
            )
            if (
                row.get('serial_no') is not None
                and not row.get('coverage_not_applicable')
            ):
                ws.cell(row_index, 1).number_format = '"第 "0" 发"'
                ws.cell(row_index, 1).comment = Comment(
                    f"所属发次：第 {row['serial_no']} 发",
                    '合同生成工具',
                )
            ws.cell(
                row_index,
                helper_current_col,
                _minor_to_wan(row['current_month_minor']) or 0,
            )
            ws.cell(
                row_index,
                helper_previous_col,
                _minor_to_wan(row['previous_unpaid_minor']) or 0,
            )
            ws.cell(
                row_index,
                current_payment_col,
                f'={helper_current_letter}{row_index}+'
                f'{helper_previous_letter}{row_index}',
            )
            if row['current_month_minor']:
                ws.cell(
                    row_index,
                    current_payment_col,
                ).fill = _STYLES.monthly_current_fill
                ws.cell(
                    row_index,
                    current_payment_col,
                ).font = _STYLES.monthly_total_font
            for column in [6, current_payment_col]:
                ws.cell(row_index, column).number_format = MONEY_FORMAT
            for node_index, node in enumerate(row['nodes']):
                amount_col = 7 + node_index * 2
                condition_node_col = amount_col + 1
                ws.cell(row_index, amount_col).number_format = MONEY_FORMAT
                if node['is_paid']:
                    ws.cell(row_index, amount_col).fill = _STYLES.monthly_paid_fill
                    ws.cell(
                        row_index,
                        condition_node_col,
                    ).fill = _STYLES.monthly_paid_fill
                    ws.cell(row_index, amount_col).font = _STYLES.monthly_total_font
                    ws.cell(
                        row_index,
                        condition_node_col,
                    ).font = _STYLES.monthly_total_font
                elif node['is_current']:
                    ws.cell(
                        row_index,
                        amount_col,
                    ).fill = _STYLES.monthly_current_fill
                    ws.cell(
                        row_index,
                        condition_node_col,
                    ).fill = _STYLES.monthly_current_fill
                    ws.cell(row_index, amount_col).font = _STYLES.monthly_total_font
                    ws.cell(
                        row_index,
                        condition_node_col,
                    ).font = _STYLES.monthly_total_font
            if row['previous_unpaid_minor']:
                ws.cell(
                    row_index,
                    current_payment_col,
                ).font = _STYLES.monthly_previous_font

        previous_total_row = reserved_data_end + 1
        current_total_row = previous_total_row + 1
        planned_total_row = previous_total_row + 2
        for row_index in (previous_total_row, current_total_row, planned_total_row):
            for column in range(1, last_col + 1):
                cell = ws.cell(row_index, column)
                cell.border = _STYLES.thin_border
                cell.fill = (
                    _STYLES.monthly_total_fill
                    if row_index == planned_total_row
                    else _STYLES.monthly_helper_fill
                )
                cell.font = _STYLES.monthly_total_font
        ws.cell(
            previous_total_row,
            current_payment_col,
            f'=SUM({helper_previous_letter}{data_start}:'
            f'{helper_previous_letter}{reserved_data_end})',
        )
        ws.cell(
            current_total_row,
            current_payment_col,
            f'=SUM({helper_current_letter}{data_start}:'
            f'{helper_current_letter}{reserved_data_end})',
        )
        ws.cell(
            planned_total_row,
            current_payment_col,
            f'=SUM({current_payment_letter}{previous_total_row}:'
            f'{current_payment_letter}{current_total_row})',
        )
        for row_index in (previous_total_row, current_total_row, planned_total_row):
            ws.cell(row_index, current_payment_col).number_format = MONEY_FORMAT
        ws.cell(planned_total_row, overdue_col, '可用银承支付金额')
        ws.cell(planned_total_row, overdue_col).alignment = _STYLES.center_alignment
        ws.cell(
            planned_total_row,
            bank_col,
            f'=SUMIF({bank_letter}{data_start}:{bank_letter}{reserved_data_end},'
            f'"<>",{current_payment_letter}{data_start}:'
            f'{current_payment_letter}{reserved_data_end})',
        )
        ws.cell(planned_total_row, bank_col).number_format = MONEY_FORMAT

        legend_paid_row = planned_total_row + 2
        legend_current_row = legend_paid_row + 1
        ws.cell(legend_paid_row, 1, '绿色的表示已付款')
        ws.cell(legend_paid_row, 1).fill = _STYLES.monthly_paid_fill
        ws.cell(legend_paid_row, 1).font = _STYLES.monthly_total_font
        ws.cell(
            legend_current_row,
            1,
            f'黄色表示{display_month}计划付款',
        )
        ws.cell(legend_current_row, 1).fill = _STYLES.monthly_current_fill
        ws.cell(legend_current_row, 1).font = _STYLES.monthly_total_font

        _STYLES.apply_column_widths(ws, columns)
        _style_monthly_sheet(
            ws,
            last_col,
            data_start,
            reserved_data_end,
            planned_total_row,
        )
        ws.print_area = (
            f'A1:{get_column_letter(last_col)}{legend_current_row}'
        )
        escaped_name = sheet_name.replace("'", "''")
        project_refs.append({
            'name': project['project_name'],
            'subsystem_name': project['subsystem_name'],
            'sheet_name': sheet_name,
            'sheet_formula_name': escaped_name,
            'current_cell': (
                f'{current_payment_letter}{current_total_row}'
            ),
            'previous_cell': (
                f'{current_payment_letter}{previous_total_row}'
            ),
            'planned_cell': (
                f'{current_payment_letter}{planned_total_row}'
            ),
            'bank_cell': (
                f'{get_column_letter(bank_col)}{planned_total_row}'
            ),
        })

    summary_columns = monthly_summary_columns(f'{int(month)}月')
    summary_headers = column_headers(summary_columns)
    summary_ws.sheet_view.showGridLines = False
    summary_ws.sheet_format.defaultRowHeight = 20
    summary_ws.column_dimensions['A'].width = 9
    for column, definition in enumerate(summary_columns, 2):
        summary_ws.column_dimensions[
            get_column_letter(column)
        ].width = definition.width
    for column, header in enumerate(summary_headers, 2):
        cell = summary_ws.cell(2, column, header)
        cell.font = _STYLES.monthly_summary_header_font
        cell.fill = _STYLES.monthly_header_fill
        cell.border = _STYLES.thin_border
        cell.alignment = _STYLES.center_alignment

    project_data_end = 2 + len(project_refs)
    reserved_project_end = max(6, project_data_end)
    for row_index in range(3, reserved_project_end + 1):
        for column in range(2, 9):
            cell = summary_ws.cell(row_index, column)
            cell.font = _STYLES.monthly_summary_body_font
            cell.border = _STYLES.thin_border
            if column in (4, 5, 6, 7):
                cell.number_format = MONEY_FORMAT
                cell.alignment = _STYLES.right_alignment

    for row_index, ref in enumerate(project_refs, 3):
        summary_ws.cell(
            row_index,
            2,
            safe_spreadsheet_value(ref['name']),
        )
        summary_ws.cell(row_index, 2).font = _STYLES.monthly_summary_bold_font
        summary_ws.cell(
            row_index,
            3,
            safe_spreadsheet_value(ref['subsystem_name']),
        )
        summary_ws.cell(row_index, 3).alignment = _STYLES.left_wrap_alignment
        summary_ws.cell(
            row_index,
            2,
        ).alignment = _STYLES.left_wrap_alignment
        summary_ws.row_dimensions[row_index].height = (
            _monthly_summary_row_height(
                ref['name'], ref['subsystem_name']
            )
        )
        summary_ws.cell(
            row_index,
            4,
            f'=SUM(E{row_index}:F{row_index})',
        )
        summary_ws.cell(
            row_index, 5,
            f"='{ref['sheet_formula_name']}'!{ref['current_cell']}",
        )
        summary_ws.cell(
            row_index, 6,
            f"='{ref['sheet_formula_name']}'!{ref['previous_cell']}",
        )
        summary_ws.cell(
            row_index, 7,
            f"='{ref['sheet_formula_name']}'!{ref['bank_cell']}",
        )
        summary_ws.cell(row_index, 8, '')

    total_row = reserved_project_end + 2
    summary_ws.cell(total_row, 2, '合计：')
    summary_ws.cell(total_row, 2).font = _STYLES.monthly_summary_bold_font
    for column in range(2, 9):
        cell = summary_ws.cell(total_row, column)
        cell.border = _STYLES.thin_border
        cell.font = _STYLES.monthly_summary_bold_font
    for column in range(5, 8):
        if project_refs:
            letter = get_column_letter(column)
            summary_ws.cell(
                total_row, column,
                f'=SUM({letter}3:{letter}{reserved_project_end})',
            )
        else:
            summary_ws.cell(total_row, column, 0)
        summary_ws.cell(total_row, column).number_format = MONEY_FORMAT
    summary_ws.cell(total_row, 4, f'=SUM(E{total_row}:F{total_row})')
    summary_ws.cell(total_row, 4).number_format = MONEY_FORMAT

    diagnostics = report.get('diagnostics') or {}
    messages = []
    labels = [
        ('unassigned_serial_count', '待补发次的付款计划'),
        ('missing_due_date_count', '缺少应付日期的付款计划'),
        ('missing_due_amount_count', '缺少应付金额的付款计划'),
        ('missing_serial_amount_count', '缺少本发金额的报表行'),
    ]
    for key, label in labels:
        if diagnostics.get(key):
            messages.append(f"{label}：{diagnostics[key]}项")
    note_row = max(12, total_row + 4)
    summary_ws.cell(
        note_row,
        2,
        '说明：本月计划付款合计包括“本月新增计划付款”和'
        '“以前月度已做付款计划未付款”预计在本月支付的金额。',
    )
    summary_ws.cell(note_row, 2).font = _STYLES.monthly_summary_bold_font
    if messages:
        summary_ws.cell(note_row, 2).comment = Comment(
            '数据检查：' + '；'.join(messages),
            '合同生成工具',
        )
    summary_ws.print_area = f'B2:H{note_row}'

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
        wb.save(path)
    finally:
        wb.close()
    return path


def _export_contracts_streaming(path, contracts, title):
    """Write a large contract iterable without retaining worksheet cells."""
    columns = CONTRACT_COLUMNS
    headers = column_headers(columns)
    status_labels = {
        'draft': '草稿', 'signed': '已签订', 'active': '履行中',
        'completed': '已完成', 'void': '已作废',
    }
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=title[:31])
    _STYLES.apply_column_widths(ws, columns)

    title_cells = [
        WriteOnlyCell(ws, value=safe_spreadsheet_value(title))
    ] + [WriteOnlyCell(ws) for _ in headers[1:]]
    title_cells[0].font = _STYLES.title_font
    title_cells[0].alignment = _STYLES.title_alignment
    ws.append(title_cells)
    ws.append([None] * len(headers))
    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = _STYLES.header_font
        cell.fill = _STYLES.header_fill
        cell.alignment = _STYLES.center_alignment
        cell.border = _STYLES.thin_border
        header_cells.append(cell)
    ws.append(header_cells)

    total_amount = 0
    row_count = 0
    for row_count, contract in enumerate(contracts, 1):
        amount = _num(contract.get('amount'))
        total_amount += amount
        values = [
            row_count,
            contract.get('project_name') or '',
            contract.get('subsystem_name') or '',
            (f"第 {contract.get('coverage_start')}–{contract.get('coverage_end')} 发"
             if contract.get('coverage_start') is not None
             and contract.get('coverage_end') is not None
             else '不适用' if contract.get('coverage_not_applicable') else '待补发次'),
            contract.get('contract_no') or '',
            contract.get('title') or '',
            contract.get('counterparty') or '',
            amount,
            contract.get('sign_date') or '',
            contract.get('owner') or '',
            status_labels.get(contract.get('status'), contract.get('status', '')),
            (contract.get('created_at') or '')[:10],
            contract.get('plan_count', 0),
        ]
        cells = []
        for ci, value in enumerate(values, 1):
            cell = WriteOnlyCell(ws, value=safe_spreadsheet_value(value))
            cell.font = _STYLES.normal_font
            cell.border = _STYLES.thin_border
            if ci == 8:
                cell.number_format = MONEY_FORMAT
            cells.append(cell)
        ws.append(cells)

    summary = ['合计' if row_count else '（无数据）', None, None, None, None, None,
               None, round(total_amount, 2) if row_count else None,
               None, None, None, None, None]
    summary_cells = []
    for ci, value in enumerate(summary, 1):
        cell = WriteOnlyCell(ws, value=value)
        cell.border = _STYLES.thin_border
        if ci == 1:
            cell.font = _STYLES.header_font
        if ci == 8:
            cell.number_format = MONEY_FORMAT
        summary_cells.append(cell)
    ws.append(summary_cells)
    try:
        wb.save(path)
    finally:
        wb.close()
    return path


def export_contracts(path, contracts, title='合同台账', streaming=False):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    if streaming:
        return _export_contracts_streaming(path, contracts, title)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    status_labels = {
        'draft': '草稿', 'signed': '已签订', 'active': '履行中',
        'completed': '已完成', 'void': '已作废',
    }
    columns = CONTRACT_COLUMNS
    headers = column_headers(columns)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(
        row=1, column=1, value=safe_spreadsheet_value(title)
    )
    title_cell.font = _STYLES.title_font
    title_cell.alignment = _STYLES.title_alignment

    # 表头（第3行）
    for ci, header in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=header)
    _apply_header_style(ws, 3, len(headers))

    # 数据行
    total_amount = 0
    for idx, c in enumerate(contracts, 1):
        ri = idx + 3
        amount = _num(c.get('amount'))
        total_amount += amount
        values = [
            idx,
            c.get('project_name') or '',
            c.get('subsystem_name') or '',
            (f"第 {c.get('coverage_start')}–{c.get('coverage_end')} 发"
             if c.get('coverage_start') is not None and c.get('coverage_end') is not None
             else '不适用' if c.get('coverage_not_applicable') else '待补发次'),
            c.get('contract_no') or '',
            c.get('title') or '',
            c.get('counterparty') or '',
            amount,
            c.get('sign_date') or '',
            c.get('owner') or '',
            status_labels.get(c.get('status'), c.get('status', '')),
            (c.get('created_at') or '')[:10],
            c.get('plan_count', 0),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=safe_spreadsheet_value(val))
            cell.font = _STYLES.normal_font
            cell.border = _STYLES.thin_border
            if ci == 8:
                cell.number_format = MONEY_FORMAT

    # 合计行
    summary_row = len(contracts) + 4
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)
    if contracts:
        ws.cell(row=summary_row, column=1, value='合计').font = _STYLES.header_font
        ws.cell(row=summary_row, column=8, value=round(total_amount, 2)).number_format = MONEY_FORMAT
    else:
        ws.cell(row=summary_row, column=1, value='（无数据）').font = _STYLES.header_font
    for ci in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=ci).border = _STYLES.thin_border

    _STYLES.apply_column_widths(ws, columns)

    try:
        wb.save(path)
    finally:
        wb.close()
    return path


def _write_handover_table(wb, sheet_name, title, columns, rows):
    headers = column_headers(columns)
    money_cols = money_column_numbers(columns)
    ws = wb.create_sheet(sheet_name[:31])
    col_count = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(
        row=1, column=1, value=safe_spreadsheet_value(title)
    )
    title_cell.font = _STYLES.title_font
    title_cell.alignment = _STYLES.title_alignment

    for ci, header in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=header)
    _apply_header_style(ws, 3, col_count)

    if rows:
        for ri, row in enumerate(rows, 4):
            for ci, value in enumerate(row, 1):
                cell = ws.cell(
                    row=ri, column=ci, value=safe_spreadsheet_value(value)
                )
                cell.font = _STYLES.normal_font
                cell.alignment = _STYLES.top_wrap_alignment
                cell.border = _STYLES.thin_border
                if ci in money_cols:
                    cell.number_format = MONEY_FORMAT
    else:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=col_count)
        cell = ws.cell(row=4, column=1, value='（无数据）')
        cell.font = _STYLES.normal_font
        cell.alignment = _STYLES.title_alignment
        for ci in range(1, col_count + 1):
            ws.cell(row=4, column=ci).border = _STYLES.thin_border

    last_row = max(4, len(rows) + 3)
    ws.auto_filter.ref = f'A3:{get_column_letter(col_count)}{last_row}'
    ws.freeze_panes = 'A4'
    _STYLES.apply_column_widths(ws, columns)
    return ws


def export_handover_checklist(path, data):
    """Export an employee handover checklist workbook."""
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = '交接总览'

    title = f"{data.get('owner', '')} 交接清单"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(
        row=1, column=1, value=safe_spreadsheet_value(title)
    )
    title_cell.font = _STYLES.title_font
    title_cell.alignment = _STYLES.title_alignment

    summary = data.get('summary') or {}
    overview_rows = [
        ('员工姓名', data.get('owner', '')),
        ('生成时间', data.get('generated_at', '')),
        ('数据范围', '包含已完成/已归档' if data.get('include_closed') else '未完成/未归档'),
        ('合同数量', summary.get('contract_count', 0)),
        ('付款计划数量', summary.get('payment_count', 0)),
        ('待付款金额', summary.get('outstanding_payment_amount', 0)),
        ('采购项目数量', summary.get('project_count', 0)),
        ('待处理采购数量', summary.get('active_project_count', 0)),
        ('风险/待办数量', summary.get('risk_count', 0)),
        ('文件数量', summary.get('file_count', 0)),
    ]
    ws.cell(row=3, column=1, value=HANDOVER_OVERVIEW_COLUMNS[0].header)
    ws.cell(row=3, column=2, value=HANDOVER_OVERVIEW_COLUMNS[1].header)
    _apply_header_style(ws, 3, 2)
    for ri, row in enumerate(overview_rows, 4):
        ws.cell(row=ri, column=1, value=row[0])
        ws.cell(row=ri, column=2, value=safe_spreadsheet_value(row[1]))
        for ci in (1, 2):
            cell = ws.cell(row=ri, column=ci)
            cell.font = _STYLES.normal_font
            cell.border = _STYLES.thin_border
        if row[0] == '待付款金额':
            ws.cell(row=ri, column=2).number_format = MONEY_FORMAT
    _STYLES.apply_column_widths(ws, HANDOVER_OVERVIEW_COLUMNS)

    contract_rows = []
    for idx, row in enumerate(data.get('contracts') or [], 1):
        contract_rows.append([
            idx,
            row.get('contract_no') or '',
            row.get('title') or '',
            row.get('counterparty') or '',
            row.get('amount') or 0,
            row.get('status_label') or '',
            row.get('sign_date') or '',
            row.get('expiry_date') or '',
            row.get('project_name') or '',
            row.get('plan_count') or 0,
            row.get('unpaid_plan_count') or 0,
            row.get('unpaid_amount') or 0,
            row.get('docx_path') or '',
        ])
    _write_handover_table(
        wb,
        '合同清单',
        '合同清单',
        HANDOVER_TABLE_COLUMNS['contracts'],
        contract_rows,
    )

    payment_rows = []
    for idx, row in enumerate(data.get('payments') or [], 1):
        payment_rows.append([
            idx,
            row.get('contract_no') or '',
            row.get('contract_title') or '',
            row.get('phase_name') or '',
            row.get('due_date') or '',
            row.get('due_amount') or 0,
            row.get('paid_amount') or 0,
            row.get('unpaid_amount') or 0,
            row.get('confirm_status_label') or '',
            row.get('payment_status_label') or '',
            row.get('condition_text') or '',
            row.get('project_name') or '',
        ])
    _write_handover_table(
        wb,
        '付款计划',
        '付款计划',
        HANDOVER_TABLE_COLUMNS['payments'],
        payment_rows,
    )

    project_rows = []
    for idx, row in enumerate(data.get('projects') or [], 1):
        project_rows.append([
            idx,
            row.get('project_no') or '',
            row.get('project_name') or '',
            row.get('status_label') or '',
            row.get('method_label') or '',
            row.get('demand_department') or '',
            row.get('budget_amount') or 0,
            row.get('target_price_amount') or 0,
            row.get('item_count') or 0,
            row.get('supplier_count') or 0,
            row.get('quote_count') or 0,
            row.get('pending_clarification_count') or 0,
            row.get('updated_at') or '',
        ])
    _write_handover_table(
        wb,
        '采购项目',
        '采购项目',
        HANDOVER_TABLE_COLUMNS['projects'],
        project_rows,
    )

    risk_rows = []
    for idx, row in enumerate(data.get('risks') or [], 1):
        risk_rows.append([
            idx,
            row.get('risk_type') or '',
            row.get('related_no') or '',
            row.get('related_name') or '',
            row.get('detail') or '',
            row.get('amount') if row.get('amount') != '' else '',
            row.get('due_date') or '',
            row.get('status') or '',
        ])
    _write_handover_table(
        wb,
        '待办风险',
        '待办风险',
        HANDOVER_TABLE_COLUMNS['risks'],
        risk_rows,
    )

    file_rows = []
    for idx, row in enumerate(data.get('files') or [], 1):
        file_rows.append([
            idx,
            row.get('source') or '',
            row.get('related_no') or '',
            row.get('related_name') or '',
            row.get('file_type') or '',
            row.get('original_name') or '',
            row.get('relative_path') or '',
            row.get('size_bytes') or '',
            row.get('created_at') or '',
        ])
    _write_handover_table(
        wb,
        '文件清单',
        '文件清单',
        HANDOVER_TABLE_COLUMNS['files'],
        file_rows,
    )

    try:
        wb.save(path)
    finally:
        wb.close()
    return path
