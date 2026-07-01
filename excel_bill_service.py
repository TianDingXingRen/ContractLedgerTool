# -*- coding: utf-8 -*-
"""Excel 单据生成服务模块

提供:
- 预置单据表头定义 (4种预设)
- 从合同采购标的 (table_3) 提取数据
- 列映射与 Excel 文件生成
"""

import json
import os
import re
import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import ledger_store
import template_def
from utils.logger import get_logger

# ── 样式常量 ──
_HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=11)
_NORMAL_FONT = Font(name='Microsoft YaHei', size=10)
_HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# ── 预置单据表头定义 ──

BILL_PRESETS = {}

# ── 合同中 table_3 (采购标的) 列定义 ──
PROCUREMENT_TABLE_KEY = "table_3"
PROCUREMENT_COLUMNS = [
    {"key": "序号", "label": "序号"},
    {"key": "product_name", "label": "产品名称"},
    {"key": "spec", "label": "规格型号"},
    {"key": "uom", "label": "单位"},
    {"key": "qty", "label": "订货数量"},
    {"key": "unit_price", "label": "单价(万元)"},
    {"key": "subtotal", "label": "合计(万元)"},
    {"key": "tax_rate", "label": "增值税率"},
    {"key": "remark", "label": "备注"},
]

# ── 初始化预置 ──

def _init_presets():
    """初始化预置表头定义"""
    global BILL_PRESETS

    # 标准请购单 - 15列表头
    BILL_PRESETS["standard_pr"] = {
        "name": "标准请购单",
        "description": "适用于常规物资/服务请购场景",
        "header_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "is_executed", "label": "是否执行", "width": 10},
            {"key": "budget_type", "label": "预算类型", "width": 12},
            {"key": "estimated_amount", "label": "预估金额", "width": 14},
            {"key": "pr_type", "label": "请购类型", "width": 12},
            {"key": "dept", "label": "需求部门", "width": 14},
            {"key": "person", "label": "人员", "width": 12},
            {"key": "applicant_dept", "label": "申请人部门", "width": 14},
            {"key": "ops_category", "label": "运营管理类别", "width": 16},
            {"key": "commander", "label": "总指挥/牵头人", "width": 16},
            {"key": "commander_approval", "label": "是否总指挥/牵头人审批", "width": 22},
            {"key": "is_fixed_asset", "label": "是否科研生产类固定资产", "width": 22},
            {"key": "is_single_source", "label": "是否科研生产单一来源采购", "width": 24},
            {"key": "remark", "label": "备注", "width": 20},
            {"key": "save_action", "label": "保存动作", "width": 12},
        ],
        "detail_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "line_no", "label": "行号", "width": 8},
            {"key": "material_code", "label": "物料编码", "width": 16},
            {"key": "material_name", "label": "物料名称", "width": 28},
            {"key": "total_qty", "label": "投产总数(个)", "width": 16},
            {"key": "required_date", "label": "需求日期", "width": 14},
            {"key": "suggested_order_date", "label": "建议订货日期", "width": 16},
            {"key": "buyer", "label": "采购员", "width": 12},
            {"key": "unit_price_tax", "label": "含税单价", "width": 14},
            {"key": "total_tax", "label": "本币价税合计", "width": 16},
            {"key": "cooperation_content", "label": "合作内容", "width": 24},
            {"key": "product_no", "label": "产品号", "width": 16},
        ],
        "execution_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "exec_status", "label": "执行状态", "width": 12},
            {"key": "ncc_no", "label": "NCC单据号", "width": 20},
            {"key": "error_msg", "label": "错误信息", "width": 30},
            {"key": "screenshot_path", "label": "截图路径", "width": 30},
            {"key": "start_time", "label": "开始时间", "width": 20},
            {"key": "end_time", "label": "结束时间", "width": 20},
        ],
    }

    # 科研生产请购单 - 15列表头
    BILL_PRESETS["rd_pr"] = {
        "name": "科研生产请购单",
        "description": "适用于科研生产类物资请购，含项目编号",
        "header_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "is_executed", "label": "是否执行", "width": 10},
            {"key": "budget_type", "label": "预算类型", "width": 12},
            {"key": "estimated_amount", "label": "预估金额", "width": 14},
            {"key": "pr_type", "label": "请购类型", "width": 12},
            {"key": "dept", "label": "需求部门", "width": 14},
            {"key": "person", "label": "人员", "width": 12},
            {"key": "applicant_dept", "label": "申请人部门", "width": 14},
            {"key": "project_no", "label": "项目编号", "width": 14},
            {"key": "ops_category", "label": "运营管理类别", "width": 16},
            {"key": "commander", "label": "总指挥/牵头人", "width": 16},
            {"key": "commander_approval", "label": "是否总指挥/牵头人审批", "width": 22},
            {"key": "is_fixed_asset", "label": "是否科研生产类固定资产", "width": 22},
            {"key": "is_single_source", "label": "是否科研生产单一来源采购", "width": 24},
            {"key": "remark", "label": "备注", "width": 20},
        ],
        "detail_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "line_no", "label": "行号", "width": 8},
            {"key": "material_code", "label": "物料编码", "width": 16},
            {"key": "material_name", "label": "物料名称", "width": 28},
            {"key": "spec", "label": "规格型号", "width": 16},
            {"key": "total_qty", "label": "投产总数(个)", "width": 16},
            {"key": "required_date", "label": "需求日期", "width": 14},
            {"key": "suggested_order_date", "label": "建议订货日期", "width": 16},
            {"key": "buyer", "label": "采购员", "width": 12},
            {"key": "unit_price_tax", "label": "含税单价", "width": 14},
            {"key": "total_tax", "label": "本币价税合计", "width": 16},
            {"key": "cooperation_content", "label": "合作内容", "width": 24},
        ],
        "execution_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "exec_status", "label": "执行状态", "width": 12},
            {"key": "ncc_no", "label": "NCC单据号", "width": 20},
            {"key": "error_msg", "label": "错误信息", "width": 30},
            {"key": "screenshot_path", "label": "截图路径", "width": 30},
            {"key": "start_time", "label": "开始时间", "width": 20},
            {"key": "end_time", "label": "结束时间", "width": 20},
        ],
    }

    # 简易请购单 - 7列表头
    BILL_PRESETS["simple_pr"] = {
        "name": "简易请购单",
        "description": "精简版请购单，适用于小额/紧急采购",
        "header_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "is_executed", "label": "是否执行", "width": 10},
            {"key": "estimated_amount", "label": "预估金额", "width": 14},
            {"key": "dept", "label": "需求部门", "width": 14},
            {"key": "person", "label": "人员", "width": 12},
            {"key": "commander", "label": "审批人", "width": 12},
            {"key": "remark", "label": "备注", "width": 24},
        ],
        "detail_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "line_no", "label": "行号", "width": 8},
            {"key": "material_name", "label": "物料名称", "width": 30},
            {"key": "total_qty", "label": "数量", "width": 12},
            {"key": "unit_price_tax", "label": "含税单价", "width": 14},
            {"key": "total_tax", "label": "价税合计", "width": 16},
            {"key": "buyer", "label": "采购员", "width": 12},
            {"key": "required_date", "label": "需求日期", "width": 14},
        ],
        "execution_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "exec_status", "label": "执行状态", "width": 12},
            {"key": "ncc_no", "label": "NCC单据号", "width": 20},
            {"key": "error_msg", "label": "错误信息", "width": 30},
            {"key": "start_time", "label": "开始时间", "width": 20},
            {"key": "end_time", "label": "结束时间", "width": 20},
        ],
    }

    # 服务类请购单 - 12列表头
    BILL_PRESETS["service_pr"] = {
        "name": "服务类请购单",
        "description": "适用于技术服务/外包/咨询类请购",
        "header_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "is_executed", "label": "是否执行", "width": 10},
            {"key": "budget_type", "label": "预算类型", "width": 12},
            {"key": "estimated_amount", "label": "预估金额", "width": 14},
            {"key": "pr_type", "label": "请购类型", "width": 12},
            {"key": "dept", "label": "需求部门", "width": 14},
            {"key": "person", "label": "人员", "width": 12},
            {"key": "applicant_dept", "label": "申请人部门", "width": 14},
            {"key": "ops_category", "label": "运营管理类别", "width": 16},
            {"key": "commander", "label": "总指挥/牵头人", "width": 16},
            {"key": "is_single_source", "label": "是否单一来源采购", "width": 20},
            {"key": "remark", "label": "备注", "width": 24},
        ],
        "detail_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "line_no", "label": "行号", "width": 8},
            {"key": "material_name", "label": "服务项目名称", "width": 30},
            {"key": "cooperation_content", "label": "服务内容描述", "width": 36},
            {"key": "total_qty", "label": "数量", "width": 12},
            {"key": "unit_price_tax", "label": "含税单价", "width": 14},
            {"key": "total_tax", "label": "本币价税合计", "width": 16},
            {"key": "buyer", "label": "采购员", "width": 12},
            {"key": "required_date", "label": "需求日期", "width": 14},
        ],
        "execution_columns": [
            {"key": "bill_no", "label": "单据编号", "width": 28},
            {"key": "exec_status", "label": "执行状态", "width": 12},
            {"key": "ncc_no", "label": "NCC单据号", "width": 20},
            {"key": "error_msg", "label": "错误信息", "width": 30},
            {"key": "screenshot_path", "label": "截图路径", "width": 30},
            {"key": "start_time", "label": "开始时间", "width": 20},
            {"key": "end_time", "label": "结束时间", "width": 20},
        ],
    }


_init_presets()


# ── 公开 API ──

def get_presets():
    """返回所有预置单据表头摘要列表"""
    return [
        {
            "key": preset_key,
            "name": preset["name"],
            "description": preset.get("description", ""),
            "header_count": len(preset["header_columns"]),
            "detail_count": len(preset["detail_columns"]),
        }
        for preset_key, preset in BILL_PRESETS.items()
    ]


def get_preset(preset_key):
    """按 key 获取某个预置的完整定义"""
    if preset_key not in BILL_PRESETS:
        raise ValueError("预设表头不存在: " + preset_key)
    preset = BILL_PRESETS[preset_key]
    return {
        "key": preset_key,
        "name": preset["name"],
        "description": preset.get("description", ""),
        "header_columns": preset["header_columns"],
        "detail_columns": preset["detail_columns"],
        "execution_columns": preset.get("execution_columns", []),
    }


def _column_label(key):
    for column in PROCUREMENT_COLUMNS:
        if column['key'] == key:
            return column['label']
    return str(key)


def _template_table_def(contract):
    template_name = str(contract.get('template_name') or '').strip()
    if not template_name:
        return None
    candidates = []
    for info in template_def.list_templates():
        try:
            tpl = template_def.TemplateDef.load(info['path'])
        except Exception:
            get_logger().debug(
                'Skip unreadable template while resolving Excel bill table: %s',
                info.get('path'),
                exc_info=True,
            )
            continue
        if tpl.name != template_name:
            continue
        for field in tpl.data.get('fields', []):
            if field.get('field_type') != 'table':
                continue
            text = f"{field.get('key', '')} {field.get('label', '')}".lower()
            score = sum(
                1 for keyword in ('采购', '标的', '物料', '明细', '清单', 'item', 'order')
                if keyword in text
            )
            candidates.append((score, field))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def extract_contract_table(contract_id, contract=None, table_def=None):
    """返回合同中最适合作为业务单据来源的表格及列定义。"""
    contract = contract or ledger_store.get_contract(contract_id)
    if not contract:
        return {'table_key': '', 'rows': [], 'columns': []}
    try:
        values = json.loads(contract.get('values_json') or '{}')
    except (json.JSONDecodeError, TypeError):
        return {'table_key': '', 'rows': [], 'columns': []}

    table_def = table_def or _template_table_def(contract)
    preferred_keys = [PROCUREMENT_TABLE_KEY]
    if table_def and table_def.get('key') not in preferred_keys:
        preferred_keys.append(table_def.get('key'))

    table_candidates = []
    for key, rows in values.items():
        if not isinstance(rows, list) or not any(isinstance(row, dict) for row in rows):
            continue
        row_keys = {
            row_key for row in rows if isinstance(row, dict) for row_key in row.keys()
        }
        score = len(row_keys & {column['key'] for column in PROCUREMENT_COLUMNS})
        if key in preferred_keys:
            score += 100 - preferred_keys.index(key)
        table_candidates.append((score, key, rows, row_keys))

    if not table_candidates:
        return {'table_key': '', 'rows': [], 'columns': []}
    table_candidates.sort(key=lambda item: (item[0], len(item[2])), reverse=True)
    _, table_key, rows, row_keys = table_candidates[0]

    if table_def and table_def.get('key') == table_key:
        columns = [
            {'key': column.get('key'), 'label': column.get('label') or column.get('key')}
            for column in table_def.get('columns', []) if column.get('key')
        ]
    elif table_key == PROCUREMENT_TABLE_KEY:
        columns = [dict(column) for column in PROCUREMENT_COLUMNS]
    else:
        columns = [
            {'key': key, 'label': _column_label(key)} for key in sorted(row_keys)
        ]
    return {
        'table_key': table_key,
        'rows': [dict(row) for row in rows if isinstance(row, dict)],
        'columns': columns,
    }


def extract_table_from_contract(contract_id):
    """从指定合同的 values_json 中提取 table_3 (采购标的) 数据

    Returns:
        list[dict]: 表格行列表, 每行为 {col_key: value, ...}
        如果合同不存在或无采购标的数据, 返回空列表 []
    """
    detail = extract_contract_table(contract_id)
    table_data = detail['rows']
    if not table_data:
        return []

    col_keys = [c["key"] for c in PROCUREMENT_COLUMNS]
    normalized = []
    for row in table_data:
        if not isinstance(row, dict):
            continue
        norm_row = dict(row)
        if detail['table_key'] == PROCUREMENT_TABLE_KEY:
            for ck in col_keys:
                norm_row.setdefault(ck, "")
        normalized.append(norm_row)

    return normalized


def get_contracts_for_selection():
    """获取可用于关联的合同列表"""
    contracts = ledger_store.list_contracts(q="", status="", page=1, per_page=500)
    result = []
    table_defs = {}
    for c in contracts.get("rows", []):
        template_name = c.get('template_name') or ''
        if template_name not in table_defs:
            table_defs[template_name] = _template_table_def(c)
        table_detail = extract_contract_table(
            c['id'], contract=c, table_def=table_defs[template_name]
        )
        table_data = table_detail['rows']
        has_table = bool(table_data)
        result.append({
            "id": c["id"],
            "title": c.get("title", ""),
            "contract_no": c.get("contract_no", ""),
            "counterparty": c.get("counterparty", ""),
            "has_table3": has_table,
            "table_key": table_detail['table_key'],
            "table3_columns": table_detail['columns'],
            "item_count": len(table_data),
        })
    return result


def generate_bill_excel(preset_key, header_data, detail_rows, output_dir):
    """生成 Excel 单据文件

    Args:
        preset_key: 预置表头 key
        header_data: 单据表头数据 dict
        detail_rows: 单据明细列表 list[dict]
        output_dir: 输出目录

    Returns:
        str: 生成的 .xlsx 文件路径
    """
    preset = get_preset(preset_key)
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # Sheet 1: 单据表头
    ws_header = wb.active
    ws_header.title = "单据表头"
    _write_header_sheet(ws_header, preset["header_columns"], header_data)

    # Sheet 2: 单据明细
    ws_detail = wb.create_sheet("单据明细")
    _write_detail_sheet(ws_detail, preset["detail_columns"], detail_rows)

    # Sheet 3: 执行结果
    ws_exec = wb.create_sheet("执行结果")
    _write_execution_sheet(ws_exec, preset.get("execution_columns", []),
                           header_data.get("bill_no", ""))

    # 保存
    bill_no = header_data.get("bill_no", uuid.uuid4().hex[:12])
    safe_name = re.sub(r'[^\w一-鿿-]', '_', str(bill_no))
    filename = safe_name + "_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    return path


def _write_header_sheet(ws, columns, data):
    """写入单据表头 Sheet"""
    for ci, col_def in enumerate(columns, 1):
        cell_label = ws.cell(row=1, column=ci, value=col_def["label"])
        cell_label.font = _HEADER_FONT
        cell_label.fill = _HEADER_FILL
        cell_label.alignment = Alignment(horizontal='center', vertical='center')
        cell_label.border = _THIN_BORDER

        val = data.get(col_def["key"], "")
        cell_value = ws.cell(row=2, column=ci, value=val)
        cell_value.font = _NORMAL_FONT
        cell_value.border = _THIN_BORDER
        cell_value.alignment = Alignment(vertical='center')

    for ci, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_def.get("width", 14)


def _write_detail_sheet(ws, columns, rows):
    """写入单据明细 Sheet"""
    for ci, col_def in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_def["label"])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _THIN_BORDER

    for ri, row_data in enumerate(rows, 2):
        for ci, col_def in enumerate(columns, 1):
            val = row_data.get(col_def["key"], "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical='center')
            if col_def["key"] in ("unit_price_tax", "total_tax", "total_qty", "line_no"):
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00' if col_def["key"] != "line_no" else '0'

    for ci, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_def.get("width", 14)


def _write_execution_sheet(ws, columns, bill_no):
    """写入执行结果 Sheet (仅表头, 无数据)"""
    for ci, col_def in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_def["label"])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _THIN_BORDER

    for ci, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_def.get("width", 14)


def map_contract_items_to_detail(contract_items, column_mapping, bill_no, default_values=None):
    """将合同的采购标的行映射到单据明细行

    Args:
        contract_items: extract_table_from_contract() 返回的行列表
        column_mapping: {"目标列key": "来源列key"} 映射字典
        bill_no: 单据编号
        default_values: 所有明细行共享的默认值 dict

    Returns:
        list[dict]: 可直接传入 generate_bill_excel 的 detail_rows
    """
    defaults = default_values or {}
    detail_rows = []

    for idx, item in enumerate(contract_items):
        row = {"bill_no": bill_no, "line_no": idx + 1}
        for dk, dv in defaults.items():
            row[dk] = dv
        for target_col, source_col in column_mapping.items():
            if source_col in item:
                row[target_col] = item[source_col]
            elif target_col not in row:
                row[target_col] = ""
        detail_rows.append(row)

    return detail_rows


# ── 表头数据持久化 ──

import os as _os

_DEFAULTS_DIR = None


def configure_defaults_dir(path):
    """设置可持久化的单据默认值目录。"""
    global _DEFAULTS_DIR
    _DEFAULTS_DIR = _os.path.abspath(str(path))
    _os.makedirs(_DEFAULTS_DIR, exist_ok=True)
    return _DEFAULTS_DIR


def _get_defaults_dir():
    """获取/延迟初始化 defaults 存储目录"""
    global _DEFAULTS_DIR
    if _DEFAULTS_DIR is None:
        _DEFAULTS_DIR = _os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)), "data", "excel_bill_defaults"
        )
        _os.makedirs(_DEFAULTS_DIR, exist_ok=True)
    return _DEFAULTS_DIR


def save_header_default(preset_key, name, header_data, detail_defaults=None, column_mapping=None):
    """保存一组表头填写值

    Args:
        preset_key: 对应的预置表头 key
        name: 保存名称（用户自定义标签）
        header_data: 表头数据 dict
        detail_defaults: 明细默认值 dict (采购员/日期等)
        column_mapping: 列映射 dict
    """
    import re as _re
    ddir = _get_defaults_dir()
    # 清洗 preset_key：取 basename 并校验路径不越界
    safe_preset_key = _os.path.basename(str(preset_key or ''))
    safe_name = _re.sub(r'[^\w一-鿿-]', '_', str(name))
    filename = f"{safe_preset_key}__{safe_name}.json"
    path = _os.path.realpath(_os.path.join(ddir, filename))
    ddir_real = _os.path.realpath(ddir)
    if _os.path.commonpath([ddir_real, path]) != ddir_real:
        raise ValueError('无效的预设标识')

    record = {
        "preset_key": safe_preset_key,
        "name": name,
        "header_data": header_data,
        "detail_defaults": detail_defaults or {},
        "column_mapping": column_mapping or {},
        "saved_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(record, f, ensure_ascii=False, indent=2)
    return filename


def list_header_defaults(preset_key=None):
    """列出所有已保存的表头默认值

    Args:
        preset_key: 可选，筛选指定预置的保存记录

    Returns:
        list[dict]
    """
    ddir = _get_defaults_dir()
    results = []
    if not _os.path.isdir(ddir):
        return results

    for fname in sorted(_os.listdir(ddir), reverse=True):
        if not fname.endswith(".json"):
            continue
        path = _os.path.join(ddir, fname)
        try:
            with open(path, "r", encoding="utf-8") as f:
                record = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue

        if preset_key and record.get("preset_key") != preset_key:
            continue

        results.append({
            "filename": fname,
            "preset_key": record.get("preset_key", ""),
            "name": record.get("name", ""),
            "header_count": len(record.get("header_data", {})),
            "saved_at": record.get("saved_at", ""),
            "data": record,
        })

    return results


def load_header_default(filename):
    """加载指定保存的表头默认值"""
    ddir = _get_defaults_dir()
    filename = _os.path.basename(filename or "")
    if not filename.endswith(".json"):
        raise ValueError("无效的文件名")
    ddir_real = _os.path.realpath(ddir)
    path = _os.path.realpath(_os.path.join(ddir, filename))
    if _os.path.commonpath([ddir_real, path]) != ddir_real:
        raise ValueError("无效的文件名")
    if not _os.path.isfile(path):
        raise FileNotFoundError(f"保存记录不存在: {filename}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def delete_header_default(filename):
    """删除指定保存的表头默认值"""
    ddir = _get_defaults_dir()
    filename = _os.path.basename(filename or "")
    if not filename.endswith(".json"):
        return False
    ddir_real = _os.path.realpath(ddir)
    path = _os.path.realpath(_os.path.join(ddir, filename))
    if _os.path.commonpath([ddir_real, path]) != ddir_real:
        return False
    if _os.path.isfile(path):
        _os.remove(path)
        return True
    return False
