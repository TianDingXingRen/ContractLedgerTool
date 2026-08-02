"""Excel 单据生成功能 — 单元测试"""

import os
import sys
import json
import tempfile
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import excel_bill_service


class TestPresets:
    """表头预置配置测试"""

    def test_get_presets_returns_all(self):
        presets = excel_bill_service.get_presets()
        assert len(presets) >= 3
        keys = {p["key"] for p in presets}
        assert "standard_pr" in keys
        assert "simple_pr" in keys

    def test_get_presets_structure(self):
        for p in excel_bill_service.get_presets():
            assert "key" in p
            assert "name" in p
            assert "description" in p
            assert "header_count" in p
            assert "detail_count" in p
            assert p["header_count"] > 0
            assert p["detail_count"] > 0

    def test_get_preset_standard(self):
        preset = excel_bill_service.get_preset("standard_pr")
        assert preset["name"] == "标准请购单"
        assert len(preset["header_columns"]) == 15
        assert len(preset["detail_columns"]) == 12
        assert len(preset["execution_columns"]) == 7

    def test_get_preset_simple(self):
        preset = excel_bill_service.get_preset("simple_pr")
        assert len(preset["header_columns"]) == 7
        assert len(preset["detail_columns"]) == 8

    def test_get_preset_not_found(self):
        with pytest.raises(ValueError):
            excel_bill_service.get_preset("nonexistent")

    def test_all_presets_have_valid_columns(self):
        for p in excel_bill_service.get_presets():
            preset = excel_bill_service.get_preset(p["key"])
            for col in preset["header_columns"]:
                assert "key" in col
                assert "label" in col
                assert "width" in col
            for col in preset["detail_columns"]:
                assert "key" in col
                assert "label" in col


class TestColumnDefinitions:
    """采购标的列定义测试"""

    def test_procurement_columns(self):
        cols = excel_bill_service.PROCUREMENT_COLUMNS
        assert len(cols) == 9
        keys = [c["key"] for c in cols]
        assert "product_name" in keys
        assert "qty" in keys
        assert "unit_price" in keys
        assert "subtotal" in keys

    def test_procurement_table_key(self):
        assert excel_bill_service.PROCUREMENT_TABLE_KEY == "table_3"


class TestExtractTableFromContract:
    """采购标的数据提取测试"""

    def test_extract_empty_values(self, monkeypatch):
        called_with = None

        def fake_get_contract(cid):
            nonlocal called_with
            called_with = cid
            return {"id": cid, "values_json": "{}"}

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(1)
        assert result == []
        assert called_with == 1

    def test_extract_with_table_data(self, monkeypatch):
        table_data = [
            {"序号": "1", "product_name": "测试产品A", "qty": 10, "unit_price": 100, "subtotal": 1000},
            {"序号": "2", "product_name": "测试产品B", "qty": 20, "unit_price": 200, "subtotal": 4000},
        ]

        def fake_get_contract(cid):
            return {
                "id": cid,
                "values_json": json.dumps({"table_3": table_data}),
            }

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(1)
        assert len(result) == 2
        assert result[0]["product_name"] == "测试产品A"
        assert result[0]["qty"] == 10

    def test_extract_normalizes_missing_cols(self, monkeypatch):
        def fake_get_contract(cid):
            return {
                "id": cid,
                "values_json": json.dumps({"table_3": [{"product_name": "X"}]}),
            }

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(1)
        assert len(result) == 1
        # 所有列都存在（缺失的补空字符串）
        for col in excel_bill_service.PROCUREMENT_COLUMNS:
            assert col["key"] in result[0]

    def test_extract_contract_not_found(self, monkeypatch):
        def fake_get_contract(cid):
            return None

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(999)
        assert result == []

    def test_extract_table_not_in_values(self, monkeypatch):
        def fake_get_contract(cid):
            return {"id": cid, "values_json": json.dumps({"other_field": "value"})}

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(1)
        assert result == []

    def test_extract_empty_table_array(self, monkeypatch):
        def fake_get_contract(cid):
            return {"id": cid, "values_json": json.dumps({"table_3": []})}

        monkeypatch.setattr(excel_bill_service.ledger_store, "get_contract", fake_get_contract)
        result = excel_bill_service.extract_table_from_contract(1)
        assert result == []


class TestMapContractItems:
    """列映射测试"""

    def test_basic_mapping(self):
        items = [
            {"product_name": "产品A", "qty": 10, "unit_price": 100},
            {"product_name": "产品B", "qty": 20, "unit_price": 200},
        ]
        mapping = {"material_name": "product_name", "total_qty": "qty", "unit_price_tax": "unit_price"}
        result = excel_bill_service.map_contract_items_to_detail(items, mapping, "PR-001")

        assert len(result) == 2
        assert result[0]["bill_no"] == "PR-001"
        assert result[0]["line_no"] == 1
        assert result[0]["material_name"] == "产品A"
        assert result[0]["total_qty"] == 10
        assert result[0]["unit_price_tax"] == 100

        assert result[1]["bill_no"] == "PR-001"
        assert result[1]["line_no"] == 2
        assert result[1]["material_name"] == "产品B"

    def test_mapping_with_defaults(self):
        items = [{"product_name": "X"}]
        mapping = {"material_name": "product_name"}
        defaults = {"buyer": "张三", "required_date": "2026-06-22"}
        result = excel_bill_service.map_contract_items_to_detail(items, mapping, "PR", defaults)

        assert result[0]["buyer"] == "张三"
        assert result[0]["required_date"] == "2026-06-22"
        assert result[0]["material_name"] == "X"

    def test_mapping_missing_source_col(self):
        items = [{"product_name": "X"}]
        mapping = {"material_name": "product_name", "total_qty": "nonexistent_col"}
        result = excel_bill_service.map_contract_items_to_detail(items, mapping, "PR")

        assert result[0]["total_qty"] == ""

    def test_mapping_empty_items(self):
        result = excel_bill_service.map_contract_items_to_detail([], {}, "PR")
        assert result == []


class TestGenerateExcel:
    """Excel 生成测试"""

    def test_generate_basic_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            header = {
                "bill_no": "TEST-PR-001",
                "dept": "测试部门",
                "person": "测试人员",
                "estimated_amount": 50000,
            }
            details = [
                {"bill_no": "TEST-PR-001", "line_no": 1, "material_name": "物料A", "total_qty": 10, "unit_price_tax": 100, "total_tax": 1000},
                {"bill_no": "TEST-PR-001", "line_no": 2, "material_name": "物料B", "total_qty": 5, "unit_price_tax": 200, "total_tax": 1000},
            ]
            path = excel_bill_service.generate_bill_excel("simple_pr", header, details, tmp)
            assert os.path.exists(path)
            assert path.endswith(".xlsx")
            assert os.path.getsize(path) > 0

    def test_generate_all_sheets_present(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = excel_bill_service.generate_bill_excel("standard_pr", {"bill_no": "T001"}, [], tmp)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "单据表头" in wb.sheetnames
            assert "单据明细" in wb.sheetnames
            assert "执行结果" in wb.sheetnames

    def test_generate_detail_rows_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            details = [
                {"bill_no": "T", "line_no": i, "material_name": f"Item{i}"}
                for i in range(1, 51)
            ]
            path = excel_bill_service.generate_bill_excel("standard_pr", {"bill_no": "T"}, details, tmp)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["单据明细"]
            # 标题行 + 50 行数据 = 51 行
            assert ws.max_row == 51

    def test_generate_header_data_present(self):
        with tempfile.TemporaryDirectory() as tmp:
            header = {
                "bill_no": "PR-001",
                "dept": "供应链部",
                "person": "佟翔宇",
                "is_executed": "Y",
            }
            path = excel_bill_service.generate_bill_excel("standard_pr", header, [], tmp)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["单据表头"]
            assert ws.cell(row=2, column=1).value == "PR-001"

    def test_generate_invalid_preset(self):
        with tempfile.TemporaryDirectory() as tmp:
            with pytest.raises(ValueError):
                excel_bill_service.generate_bill_excel("invalid", {}, [], tmp)


class TestIntegration:
    """集成测试 (需要 Flask app)"""

    def test_page_accessible(self, client):
        resp = client.get("/excel-bill")
        assert resp.status_code == 200

    def test_api_presets(self, client):
        resp = client.get("/api/excel-bill/presets")
        assert resp.status_code == 200
        data = resp.get_json()
        assert "presets" in data
        assert len(data["presets"]) >= 3

    def test_api_preset_detail(self, client):
        resp = client.get("/api/excel-bill/presets/standard_pr")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["name"] == "标准请购单"
        assert "header_columns" in data
        assert len(data["header_columns"]) == 15

    def test_api_preset_not_found(self, client):
        resp = client.get("/api/excel-bill/presets/nonexistent")
        assert resp.status_code == 404

    def test_api_contracts(self, client):
        resp = client.get("/api/excel-bill/contracts")
        assert resp.status_code == 200
        data = resp.get_json()
        assert "contracts" in data

    def test_generate_without_csrf_returns_400(self, client):
        resp = client.post("/excel-bill/generate", data={
            "preset_key": "simple_pr",
            "header_data": "{}",
            "detail_data": "[]",
        })
        assert resp.status_code == 400

    def test_generate_with_csrf(self, client):
        # 先获取页面以拿到 CSRF token
        page = client.get("/excel-bill")
        page.close()
        # 从 session 中提取 csrf_token
        with client.session_transaction() as sess:
            token = sess.get("_csrf_token", "")
        resp = client.post("/excel-bill/generate", data={
            "csrf_token": token,
            "preset_key": "simple_pr",
            "bill_no": "TEST-001",
            "header_data": json.dumps({"bill_no": "TEST-001", "dept": "Test"}),
            "detail_data": "[]",
        })
        assert resp.status_code == 200
        assert resp.content_type and "spreadsheet" in resp.content_type
