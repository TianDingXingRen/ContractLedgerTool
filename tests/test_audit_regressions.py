import os
import sqlite3
import subprocess
from pathlib import Path

import pytest
from openpyxl import load_workbook

import ledger_store
import procurement_store
import xlsx_exporter
from ledger_store import backups
from routes import contracts_bp
from services import (
    negotiation_service,
    procurement_file_service,
    procurement_project_service,
    quote_mapping_service,
    quote_service,
)
from utils.money import SQLITE_MAX_INTEGER, to_minor
from utils.payment_forms import payment_row_from_form


def test_numeric_inputs_reject_non_finite_and_sqlite_overflow():
    for value in ('NaN', 'Infinity', '-Infinity'):
        with pytest.raises(ValueError):
            procurement_project_service._positive_quantity(value)
        with pytest.raises(ValueError):
            negotiation_service._money_to_minor(value)

        standard_errors = []
        assert quote_service._decimal(value, '单价', standard_errors) is None
        assert standard_errors

        mapping_errors = []
        assert quote_mapping_service._decimal(
            value, '单价', mapping_errors, 2
        ) is None
        assert mapping_errors

    assert to_minor(str(SQLITE_MAX_INTEGER // 100)) <= SQLITE_MAX_INTEGER
    with pytest.raises(ValueError, match='存储范围'):
        to_minor('1e1000')


def test_payment_form_rejects_fractional_or_extreme_trigger_days():
    for value in ('1.5', '-1', '36501'):
        with pytest.raises(ValueError, match='后置天数'):
            payment_row_from_form(0, {'plan_0_trigger_days': value})


def test_contract_update_rejects_malformed_amount_without_mutation(app, client):
    contract_id = ledger_store.create_contract(
        {'title': '金额校验合同', 'amount': '123.45'}, {}, ''
    )
    with client.session_transaction() as flask_session:
        flask_session['_csrf_token'] = 'audit-token'

    response = client.post(
        f'/contracts/{contract_id}/update',
        data={
            'csrf_token': 'audit-token',
            'title': '金额校验合同',
            'status': 'draft',
            'amount': '123abc',
        },
    )

    assert response.status_code == 400
    assert ledger_store.get_contract(contract_id)['amount'] == 123.45


def test_generated_project_file_is_removed_when_registration_fails(
    tmp_path, monkeypatch
):
    base_dir = tmp_path / 'procurement'
    monkeypatch.setattr(procurement_file_service, 'BASE_DIR', base_dir)
    monkeypatch.setattr(
        procurement_store,
        'register_project_file',
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError('db failed')),
    )
    project = {'id': 1, 'project_no': 'CG-001', 'project_name': '原子写入'}

    with pytest.raises(RuntimeError, match='db failed'):
        procurement_file_service.save_generated(
            project,
            'comparison',
            'comparison.xlsx',
            lambda path: Path(path).write_bytes(b'xlsx'),
        )

    assert not list(base_dir.rglob('*.xlsx'))
    assert not list(base_dir.rglob('*.stage*'))


def test_database_backup_failure_leaves_no_partial_target(tmp_path, monkeypatch):
    source = tmp_path / 'source.db'
    target = tmp_path / 'backup.db'
    with sqlite3.connect(source) as connection:
        connection.execute('CREATE TABLE sample (id INTEGER PRIMARY KEY)')
        connection.execute('INSERT INTO sample DEFAULT VALUES')

    monkeypatch.setattr(
        backups,
        '_validate_sqlite_backup',
        lambda _path: (_ for _ in ()).throw(ValueError('invalid backup')),
    )
    with pytest.raises(ValueError, match='invalid backup'):
        backups._copy_database(str(source), str(target))

    assert not target.exists()
    assert not list(tmp_path.glob('backup.db.tmp-*'))


def test_batch_archive_reports_open_and_close_failures(tmp_path, monkeypatch):
    failures = []
    monkeypatch.setattr(
        contracts_bp.zipfile,
        'ZipFile',
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError('open failed')),
    )
    with contracts_bp._batch_archive(tmp_path / 'open.zip', failures) as archive:
        assert archive is None
    assert len(failures) == 1 and 'open failed' in str(failures[0])

    class CloseFailureArchive:
        def close(self):
            raise OSError('close failed')

    failures = []
    monkeypatch.setattr(
        contracts_bp.zipfile,
        'ZipFile',
        lambda *_args, **_kwargs: CloseFailureArchive(),
    )
    with contracts_bp._batch_archive(tmp_path / 'close.zip', failures) as archive:
        assert archive is not None
    assert len(failures) == 1 and 'close failed' in str(failures[0])


def test_standard_quote_parser_closes_workbook_on_validation_error(monkeypatch):
    class MissingSheetsWorkbook:
        sheetnames = []
        closed = False

        def close(self):
            self.closed = True

    workbook = MissingSheetsWorkbook()
    monkeypatch.setattr(quote_service, 'load_workbook', lambda *_args, **_kwargs: workbook)
    monkeypatch.setattr(
        procurement_store, 'get_project', lambda _project_id: {'project_no': 'CG-1'}
    )
    monkeypatch.setattr(
        procurement_store,
        'get_project_supplier',
        lambda _supplier_id: {'project_id': 1, 'supplier_name': '供应商'},
    )
    monkeypatch.setattr(procurement_store, 'list_project_items', lambda _project_id: [])

    _payload, errors, _warnings = quote_service.parse_standard_quote(
        'missing.xlsx', 1, 2, 1
    )

    assert errors and workbook.closed is True


def test_template_builder_assigns_stored_values_without_dynamic_html():
    source = (
        Path(__file__).resolve().parents[1]
        / 'static'
        / 'js'
        / 'template-builder.js'
    ).read_text(encoding='utf-8')
    assert 'escapeHtml' not in source
    assert 'renderTableColumns' not in source
    assert "Array.isArray(data.options) ? data.options.join('\\n')" in source
    assert 'optionsInput.value = Array.isArray(columnData.options)' in source


def test_contract_export_writes_untrusted_formula_as_text(tmp_path):
    output = tmp_path / 'contracts.xlsx'
    xlsx_exporter.export_contracts(
        output,
        [{
            'contract_no': '=HYPERLINK("https://example.invalid","open")',
            'title': '=1+1',
            'status': 'draft',
        }],
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook.active
        assert sheet.cell(4, 4).data_type == 's'
        assert sheet.cell(4, 4).value.startswith("'=")
        assert sheet.cell(4, 5).data_type == 's'
    finally:
        workbook.close()


@pytest.mark.skipif(os.name != 'nt', reason='PowerShell installer safety test')
@pytest.mark.parametrize(
    ('script', 'extra_args'),
    [
        ('install.ps1', ['-NoStart', '-NoAutostart', '-NoDesktopShortcut']),
        (
            'installer_assets/install.ps1',
            ['-NoStart', '-NoAutostart', '-NoDesktopShortcut'],
        ),
        (
            'installer_assets/uninstall.ps1',
            ['-NoPrompt', '-SkipSystemIntegrationCleanup'],
        ),
    ],
)
def test_installers_refuse_home_directory(script, extra_args):
    root = Path(__file__).resolve().parents[1]
    completed = subprocess.run(
        [
            'powershell',
            '-NoProfile',
            '-ExecutionPolicy',
            'Bypass',
            '-File',
            str(root / script),
            '-InstallDir',
            str(Path.home()),
            *extra_args,
        ],
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=30,
        check=False,
    )

    assert completed.returncode != 0
    assert 'unsafe directory' in (completed.stdout + completed.stderr)
