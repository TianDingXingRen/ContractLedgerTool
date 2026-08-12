# -*- coding: utf-8 -*-
"""Regression tests for diagnostics, backups, and editor UI affordances."""

import os
import io
import json
import tempfile
import unittest
import zipfile
from unittest import mock

import app
import ledger_store
import procurement_store
import template_def
from flask import request
from services.handover_archive import (
    MAX_FULL_PACKAGE_UPLOAD_REQUEST_BYTES,
)
from utils.session_store import save_session_data


class OperationsUiTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.original_base_dir = app.BASE_DIR
        self.original_resource_dir = app.RESOURCE_DIR
        app.reset_runtime()
        self.test_app = app.create_app(
            runtime_base_dir=self.tmp.name,
            resource_dir=self.original_resource_dir,
            run_maintenance=False,
            testing=True,
        )
        self.paths = self.test_app.extensions['runtime_paths']

    def tearDown(self):
        app.reset_runtime()
        self.tmp.cleanup()
        app.configure_runtime_paths(
            self.original_base_dir,
            self.original_resource_dir,
        )

    def _client_with_csrf(self):
        client = self.test_app.test_client()
        with client.session_transaction() as sess:
            sess['_csrf_token'] = 'test-token'
        return client

    def test_backup_create_download_restore_routes(self):
        first_docx = os.path.join(self.tmp.name, 'first.docx')
        second_docx = os.path.join(self.tmp.name, 'second.docx')
        open(first_docx, 'wb').close()
        open(second_docx, 'wb').close()

        ledger_store.create_contract({'title': '第一份合同'}, {}, first_docx)
        with self._client_with_csrf() as client:
            response = client.post(
                '/backups/create',
                data={'csrf_token': 'test-token'},
                headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
            )
            try:
                self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
                payload = response.get_json()
            finally:
                response.close()

            filename = payload['backup']['filename']
            self.assertTrue(filename.endswith('.db'))
            self.assertTrue(os.path.exists(os.path.join(ledger_store.BACKUP_DIR, filename)))

            ledger_store.create_contract({'title': '第二份合同'}, {}, second_docx)
            self.assertEqual(ledger_store.get_contract_stats()['total'], 2)

            restore_response = client.post(
                f'/backups/{filename}/restore',
                data={'csrf_token': 'test-token'},
                headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
            )
            try:
                self.assertEqual(restore_response.status_code, 200, restore_response.get_data(as_text=True))
            finally:
                restore_response.close()

            self.assertEqual(ledger_store.get_contract_stats()['total'], 1)

            download_response = client.get(f'/backups/{filename}/download')
            try:
                self.assertEqual(download_response.status_code, 200)
                self.assertGreater(len(download_response.get_data()), 0)
            finally:
                download_response.close()

    def test_backup_download_rejects_unknown_file(self):
        with self.test_app.test_client() as client:
            response = client.get('/backups/not-a-backup.txt/download')
            try:
                self.assertEqual(response.status_code, 404)
            finally:
                response.close()

    def test_full_backup_package_create_download_restore_routes(self):
        upload_path = os.path.join(self.paths.uploads_dir, 'source.docx')
        template_path = os.path.join(template_def.TEMPLATES_DIR, 'handover.contract-template')
        output_path = os.path.join(self.paths.output_dir, 'generated.docx')
        defaults_dir = os.path.join(ledger_store.DATA_DIR, 'excel_bill_defaults')
        invoice_files_dir = os.path.join(ledger_store.DATA_DIR, 'invoice_files', '1')
        os.makedirs(defaults_dir, exist_ok=True)
        os.makedirs(invoice_files_dir, exist_ok=True)
        with open(upload_path, 'wb') as f:
            f.write(b'upload')
        with open(template_path, 'w', encoding='utf-8') as f:
            f.write('{}')
        with open(output_path, 'wb') as f:
            f.write(b'output')
        with open(os.path.join(defaults_dir, 'preset.json'), 'w', encoding='utf-8') as f:
            json.dump({'name': 'preset'}, f)
        invoice_attachment = os.path.join(invoice_files_dir, 'invoice.pdf')
        with open(invoice_attachment, 'wb') as f:
            f.write(b'invoice-attachment')

        ledger_store.create_contract({'title': '第一份合同', 'owner': '张三'}, {}, output_path)
        with self._client_with_csrf() as client:
            response = client.post(
                '/backups/full/create',
                data={'csrf_token': 'test-token'},
                headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
            )
            try:
                self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
                payload = response.get_json()
            finally:
                response.close()

            filename = payload['package']['filename']
            package_path = os.path.join(ledger_store.BACKUP_DIR, 'packages', filename)
            self.assertTrue(os.path.isfile(package_path))
            with zipfile.ZipFile(package_path) as zf:
                names = set(zf.namelist())
            self.assertIn('manifest.json', names)
            self.assertIn('data/contracts.db', names)
            self.assertIn('config.json', names)
            self.assertIn('uploads/source.docx', names)
            self.assertIn('templates/handover.contract-template', names)
            self.assertIn('output/generated.docx', names)
            self.assertIn('data/invoice_files/1/invoice.pdf', names)
            self.assertFalse(any(name.startswith('sessions/') for name in names))
            self.assertFalse(any(name.startswith('logs/') for name in names))
            self.assertFalse(any(name.startswith('data/backups/') for name in names))

            download = client.get(f'/backups/full/{filename}/download')
            try:
                self.assertEqual(download.status_code, 200)
                self.assertGreater(len(download.get_data()), 0)
            finally:
                download.close()

            ledger_store.create_contract({'title': '第二份合同', 'owner': '张三'}, {}, '')
            os.remove(output_path)
            os.remove(invoice_attachment)
            self.assertEqual(ledger_store.get_contract_stats()['total'], 2)

            restore = client.post(
                f'/backups/full/{filename}/restore',
                data={'csrf_token': 'test-token'},
                headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
            )
            try:
                self.assertEqual(restore.status_code, 200, restore.get_data(as_text=True))
                restore_payload = restore.get_json()
            finally:
                restore.close()

            self.assertEqual(ledger_store.get_contract_stats()['total'], 1)
            self.assertTrue(os.path.exists(output_path))
            self.assertTrue(os.path.exists(invoice_attachment))
            self.assertTrue(
                restore_payload['rollback']['filename'].startswith('handover_full_')
            )

    def test_full_backup_upload_rejects_non_app_zip(self):
        bad_zip = io.BytesIO()
        with zipfile.ZipFile(bad_zip, 'w') as zf:
            zf.writestr('notes.txt', 'not a handover package')
        bad_zip.seek(0)

        with self._client_with_csrf() as client:
            response = client.post(
                '/backups/full/upload',
                data={'csrf_token': 'test-token', 'file': (bad_zip, 'bad.zip')},
                content_type='multipart/form-data',
                headers={
                    'Accept': 'application/json',
                    'X-Requested-With': 'XMLHttpRequest',
                    'X-CSRF-Token': 'test-token',
                },
            )
            try:
                self.assertEqual(response.status_code, 400)
                self.assertFalse(response.get_json()['success'])
            finally:
                response.close()

    def test_full_backup_upload_uses_dedicated_request_limit(self):
        self.test_app.config['MAX_CONTENT_LENGTH'] = 256
        observed = {}

        def capture_limit(file_storage, paths=None):
            observed['limit'] = request.max_content_length
            raise ValueError('stop after request parsing')

        payload = io.BytesIO(b'x' * 1024)
        with mock.patch(
            'routes.settings_bp.handover_service.upload_full_backup_package',
            side_effect=capture_limit,
        ), self._client_with_csrf() as client:
            response = client.post(
                '/backups/full/upload',
                data={
                    'csrf_token': 'test-token',
                    'file': (payload, 'backup.zip'),
                },
                content_type='multipart/form-data',
                headers={'X-CSRF-Token': 'test-token'},
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            observed['limit'], MAX_FULL_PACKAGE_UPLOAD_REQUEST_BYTES
        )

    def test_full_backup_upload_rejects_missing_header_before_body_parse(self):
        with mock.patch(
            'flask.wrappers.Request.form',
            new_callable=mock.PropertyMock,
            side_effect=AssertionError('multipart body was parsed'),
        ), mock.patch(
            'routes.settings_bp.handover_service.upload_full_backup_package'
        ) as upload, self._client_with_csrf() as client:
            response = client.post(
                '/backups/full/upload',
                data=b'unparsed multipart bytes',
                content_type='multipart/form-data; boundary=test',
            )

        self.assertEqual(response.status_code, 400)
        upload.assert_not_called()

    def test_full_backup_upload_enforces_dedicated_request_limit(self):
        self.test_app.config['FULL_BACKUP_MAX_CONTENT_LENGTH'] = 256
        with self._client_with_csrf() as client:
            response = client.post(
                '/backups/full/upload',
                data={'file': (io.BytesIO(b'x' * 1024), 'backup.zip')},
                content_type='multipart/form-data',
                headers={'X-CSRF-Token': 'test-token'},
            )

        self.assertEqual(response.status_code, 413)

    def test_full_backup_creation_discards_package_that_fails_self_check(self):
        packages_dir = os.path.join(ledger_store.BACKUP_DIR, 'packages')
        before = set(os.listdir(packages_dir)) if os.path.isdir(packages_dir) else set()
        with mock.patch(
            'services.handover_service.validate_full_backup_package',
            side_effect=ValueError('self-check failed'),
        ):
            with self.assertRaisesRegex(ValueError, 'self-check failed'):
                from services import handover_service
                handover_service.create_full_backup_package(paths=self.paths)

        self.assertEqual(set(os.listdir(packages_dir)), before)

    def test_handover_export_contains_contract_payment_procurement_and_risks(self):
        from openpyxl import load_workbook

        docx_path = os.path.join(
            self.paths.output_dir,
            'handover-contract.docx',
        )
        with open(docx_path, 'wb') as f:
            f.write(b'docx')
        ledger_store.create_contract_with_plans(
            {
                'contract_no': 'HT-001',
                'title': '交接合同',
                'counterparty': '乙方单位',
                'amount': 1000,
                'owner': '李四',
                'status': 'active',
            },
            {},
            docx_path,
            [{
                'phase_name': '首付款',
                'due_date': '2026-01-01',
                'due_amount': 1000,
                'paid_amount': 200,
                'paid_date': '2026-01-02',
                'confirm_status': 'confirmed',
            }],
        )
        project_id = procurement_store.create_project({
            'project_no': 'CG-001',
            'project_name': '交接采购项目',
            'owner': '李四',
            'demand_department': '需求部门',
            'budget_minor': 50000,
            'target_price_minor': 45000,
        })
        procurement_store.add_project_supplier(project_id, {'supplier_name': '供应商A'})
        procurement_store.register_project_file(
            project_id,
            'quote_template',
            'procurement/CG-001/template.xlsx',
            original_name='template.xlsx',
            size_bytes=128,
        )

        with self._client_with_csrf() as client:
            response = client.post(
                '/backups/handover/export',
                data={'owner': '李四', 'csrf_token': 'test-token'},
            )
            try:
                if response.status_code != 200:
                    self.fail(response.get_data(as_text=True))
                workbook_bytes = response.get_data()
            finally:
                response.close()

        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        self.assertEqual(
            set(wb.sheetnames),
            {'交接总览', '合同清单', '付款计划', '采购项目', '待办风险', '文件清单'},
        )
        self.assertEqual(wb['合同清单']['B4'].value, 'HT-001')
        self.assertEqual(wb['付款计划']['B4'].value, 'HT-001')
        self.assertEqual(wb['采购项目']['B4'].value, 'CG-001')
        risk_values = [
            wb['待办风险'].cell(row=row, column=2).value
            for row in range(4, wb['待办风险'].max_row + 1)
        ]
        self.assertIn('付款逾期', risk_values)
        self.assertIn('报价不完整', risk_values)

    def test_backups_page_exposes_full_package_and_handover_controls(self):
        with self.test_app.test_client() as client:
            response = client.get('/backups')
            try:
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
            finally:
                response.close()

        self.assertIn('data-testid="create-full-backup"', html)
        self.assertIn('data-testid="export-handover"', html)
        self.assertIn('/backups/full/upload', html)

    def test_diagnostics_open_folder_uses_whitelisted_paths(self):
        with self._client_with_csrf() as client:
            with mock.patch('routes.settings_bp._open_folder') as open_folder:
                response = client.post(
                    '/diagnostics/open-folder',
                    data={'folder': 'backups', 'csrf_token': 'test-token'},
                    headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
                )
                try:
                    self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
                    payload = response.get_json()
                finally:
                    response.close()

                self.assertTrue(payload['success'])
                self.assertEqual(os.path.abspath(payload['path']), os.path.abspath(ledger_store.BACKUP_DIR))
                open_folder.assert_called_once()

                invalid = client.post(
                    '/diagnostics/open-folder',
                    data={'folder': 'windows', 'csrf_token': 'test-token'},
                    headers={'Accept': 'application/json', 'X-Requested-With': 'XMLHttpRequest'},
                )
                try:
                    self.assertEqual(invalid.status_code, 400)
                    self.assertFalse(invalid.get_json()['success'])
                finally:
                    invalid.close()

    def test_diagnostics_page_exposes_one_click_actions(self):
        with self.test_app.test_client() as client:
            response = client.get('/diagnostics')
            try:
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
            finally:
                response.close()

        self.assertIn('data-testid="copy-diagnostics"', html)
        self.assertIn('data-testid="refresh-diagnostics"', html)
        self.assertIn('/diagnostics/open-folder', html)
        self.assertIn('diagnostics-folder-form', html)
        self.assertIn('/backups', html)

    def test_editor_page_exposes_sidebar_filters_and_result_panel(self):
        sid = 'editor-ui-test'
        save_session_data(sid, {
            'template_name': '测试模板',
            'step': 'editor',
            'fields': [
                {'id': 1, 'key': 'party_a', 'label': '甲方', 'field_type': 'text', 'required': True},
                {'id': 2, 'key': 'total', 'label': '合计', 'field_type': 'calculated', 'formula': '1 + 1'},
                {
                    'id': 3,
                    'key': 'items',
                    'label': '明细',
                    'field_type': 'table',
                    'columns': [{'key': 'name', 'label': '名称'}],
                    'default_rows': [],
                },
            ],
        }, self.paths)

        with self.test_app.test_client() as client:
            with client.session_transaction() as sess:
                sess['sid'] = sid
            response = client.get('/editor')
            try:
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
            finally:
                response.close()

        self.assertIn('data-testid="editor-sidebar"', html)
        self.assertIn('data-testid="editor-split-shell"', html)
        self.assertIn('data-testid="editor-assist-panel"', html)
        self.assertIn('data-testid="editor-live-preview"', html)
        self.assertIn('contract-preview-page', html)
        self.assertIn('livePreviewSummary', html)
        self.assertIn('id="contractEditorConfig"', html)
        self.assertIn('"previewBlocks"', html)
        self.assertIn('"previewWarnings"', html)
        self.assertIn('"templateRevision"', html)
        self.assertIn('"draftScope"', html)
        self.assertIn('"draftPageId"', html)
        self.assertIn('data-testid="editor-missing-fields"', html)
        self.assertIn('data-testid="editor-structure-list"', html)
        self.assertIn('id="fieldNavigator"', html)
        self.assertIn('data-filter="required"', html)
        self.assertIn('data-filter="calc"', html)
        self.assertIn('data-testid="generation-result"', html)
        self.assertIn('"fields"', html)
        self.assertIn('生成设置', html)
        self.assertIn('/static/js/editor.js', html)
        editor_script = os.path.join(app.RESOURCE_DIR, 'static', 'js', 'editor.js')
        with open(editor_script, 'r', encoding='utf-8') as f:
            editor_js = f.read()
            self.assertIn('function showGenerationResult', editor_js)
            self.assertIn('function renderLivePreview', editor_js)
            self.assertIn('function bindAssistPanel', editor_js)
        with open(os.path.join(app.RESOURCE_DIR, 'static', 'js', 'editor-draft.js'), 'r', encoding='utf-8') as f:
            editor_draft = f.read()
        self.assertIn("data['_table_' + fid]", editor_draft)
        self.assertIn("data['_table_' + fid] || data['_table_data_' + fid]", editor_draft)
        self.assertNotIn('window.location.href = result.detailUrl', html)

    def test_create_template_table_editor_uses_stacked_column_cards(self):
        with self.test_app.test_client() as client:
            response = client.get('/create-template')
            try:
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
            finally:
                response.close()

        builder_path = os.path.join(app.RESOURCE_DIR, 'static', 'js', 'template-builder.js')
        with open(builder_path, 'r', encoding='utf-8') as f:
            builder = f.read()
        self.assertIn('col-default-wrap', builder)
        self.assertIn('col-formula-wrap', builder)
        self.assertIn('refreshColumnRemoveButtons', builder)
        self.assertNotIn('max-w-2xl', html)

    def test_builder_preview_expands_with_page(self):
        with self.test_app.test_client() as client:
            response = client.get('/create-template')
            try:
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
            finally:
                response.close()

        self.assertIn('编制模板', html)

    def test_manual_template_save_preserves_table_column_types(self):
        existing_files = set(os.listdir(template_def.TEMPLATES_DIR))
        form = {
            'csrf_token': 'test-token',
            'template_name': '表格测试模板',
            'stored_name': '',
            'field_label_0': '明细表',
            'field_key_0': 'items',
            'field_type_0': 'table',
            'field_required_0': 'on',
            'col_label_0_0': '产品名称',
            'col_type_0_0': 'text',
            'col_default_0_0': '标准产品',
            'col_label_0_1': '数量',
            'col_type_0_1': 'text',
            'col_label_0_2': '单价',
            'col_type_0_2': 'text',
            'col_label_0_3': '小计',
            'col_type_0_3': 'calculated',
            'col_formula_0_3': 'qty * unit_price',
            'col_default_0_3': 'should be ignored',
        }
        with self._client_with_csrf() as client:
            response = client.post('/template/manual-save', data=form)
            try:
                self.assertEqual(response.status_code, 302, response.get_data(as_text=True))
            finally:
                response.close()

        files = set(os.listdir(template_def.TEMPLATES_DIR)) - existing_files
        self.assertEqual(len(files), 1)
        saved_file = next(iter(files))
        with open(os.path.join(template_def.TEMPLATES_DIR, saved_file), 'r', encoding='utf-8') as f:
            data = json.load(f)

        columns = data['fields'][0]['columns']
        self.assertEqual(columns[0]['default_value'], '标准产品')
        self.assertEqual(columns[1]['key'], 'qty')
        self.assertEqual(columns[2]['key'], 'unit_price')
        self.assertEqual(columns[3]['field_type'], 'calculated')
        self.assertEqual(columns[3]['formula'], 'qty * unit_price')
        self.assertEqual(columns[3]['default_value'], '')


if __name__ == '__main__':
    unittest.main()
