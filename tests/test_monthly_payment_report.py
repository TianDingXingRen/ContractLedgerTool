import io

import pytest
from openpyxl import load_workbook

import ledger_store
import xlsx_exporter


def _contract_with_serials(*, title='月报合同', project='火箭项目'):
    contract_id = ledger_store.create_contract({
        'contract_no': f'HT-{title}',
        'title': title,
        'counterparty': '乙方科技有限公司',
        'amount': 3_000_000,
        'status': 'active',
        'project_name': project,
        'coverage_start': 101,
        'coverage_end': 102,
    }, {'甲方': '甲方技术有限公司'}, f'/{title}.docx')
    serials = ledger_store.list_contract_serials(contract_id)
    ledger_store.save_contract_serial_amounts(contract_id, [
        {'id': serials[0]['id'], 'amount': '1200000', 'remark': ''},
        {'id': serials[1]['id'], 'amount': '1800000', 'remark': ''},
    ])
    return contract_id, ledger_store.list_contract_serials(contract_id)


def test_contract_serial_ledger_is_generated_and_validates_plan_links(tmp_db):
    contract_id, serials = _contract_with_serials()
    assert [row['serial_no'] for row in serials] == [101, 102]
    assert [row['serial_amount'] for row in serials] == [1_200_000, 1_800_000]

    plan_id = ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serials[0]['id'],
        'phase_name': '首付款',
        'due_date': '2026-05-20',
        'due_amount': 120_000,
        'confirm_status': 'confirmed',
    })
    plan = ledger_store.get_payment_plan(plan_id)
    assert plan['serial_no'] == 101
    assert plan['serial_amount'] == 1_200_000

    other_id, other_serials = _contract_with_serials(title='另一合同')
    assert other_id != contract_id
    with pytest.raises(ValueError, match='不属于当前合同'):
        ledger_store.insert_payment_plan(contract_id, {
            'contract_serial_id': other_serials[0]['id'],
            'phase_name': '错误关联',
        })

    ledger_store.update_contract(contract_id, {
        'coverage_start': 102,
        'coverage_end': 103,
    })
    all_serials = ledger_store.list_contract_serials(
        contract_id, include_inactive=True
    )
    assert [(row['serial_no'], row['status']) for row in all_serials] == [
        (101, 'inactive'), (102, 'active'), (103, 'active'),
    ]


def test_monthly_report_groups_nodes_by_contract_serial_without_notice_data(tmp_db):
    contract_id, serials = _contract_with_serials()
    serial_id = serials[0]['id']
    ledger_store.insert_payment_plans(contract_id, [
        {
            'contract_serial_id': serial_id,
            'phase_name': '首付款',
            'due_date': '2026-04-20',
            'due_amount': 300_000,
            'paid_amount': 100_000,
            'paid_date': '2026-04-21',
            'confirm_status': 'confirmed',
            'remark': '上月未付原因：审批尚未完成',
        },
        {
            'contract_serial_id': serial_id,
            'phase_name': '进度款',
            'due_date': '2026-05-20',
            'due_amount': 500_000,
            'confirm_status': 'confirmed',
            'condition_text': '完成阶段验收',
            'remark': '银承：6个月',
        },
        {
            'contract_serial_id': serial_id,
            'phase_name': '已付款节点',
            'due_date': '2026-03-20',
            'due_amount': 100_000,
            'paid_amount': 100_000,
            'paid_date': '2026-03-20',
            'confirm_status': 'confirmed',
        },
        {
            'contract_serial_id': serial_id,
            'phase_name': '尾款',
            'due_date': '2026-08-20',
            'due_amount': 300_000,
            'confirm_status': 'confirmed',
        },
        {
            'phase_name': '待补编号',
            'due_date': '2026-05-25',
            'due_amount': 10_000,
            'confirm_status': 'confirmed',
        },
        {
            'contract_serial_id': serials[1]['id'],
            'phase_name': '待确认未来款',
            'due_date': '2026-09-20',
            'due_amount': 300_000,
            'confirm_status': 'pending',
        },
    ])

    report = ledger_store.build_monthly_payment_report('2026-05')
    assert report['node_count'] == 4
    assert len(report['rows']) == 3
    row = next(item for item in report['rows'] if item['serial_no'] == 101)
    assert row['serial_no'] == 101
    assert row['serial_amount_minor'] == 120_000_000
    assert row['current_month_minor'] == 50_000_000
    assert row['previous_unpaid_minor'] == 20_000_000
    assert row['planned_payment_minor'] == 70_000_000
    assert row['party_a'] == '甲方技术有限公司'
    assert row['bank_acceptance'] == '6个月'
    assert row['prior_unpaid_reason'] == '审批尚未完成'
    unassigned = next(
        item for item in report['rows']
        if item['contract_serial_id'] is None
    )
    assert unassigned['nodes'][0]['condition'].startswith('2026-05-25')
    assert unassigned['nodes'][0]['is_current'] is True
    future = next(item for item in report['rows'] if item['serial_no'] == 102)
    assert future['planned_payment_minor'] == 0
    assert future['nodes'][0]['condition'].startswith('2026-09-20')
    assert future['nodes'][0]['is_current'] is False
    assert report['diagnostics']['unassigned_serial_count'] == 1


def test_not_applicable_contract_skips_serial_diagnostics_and_exports_label(
    tmp_db, tmp_path,
):
    contract_id = ledger_store.create_contract({
        'contract_no': 'HT-NO-SERIAL',
        'title': '非发次服务合同',
        'counterparty': '北岳质量技术有限公司',
        'project_name': '综合保障项目',
        'coverage_not_applicable': True,
        'coverage_start': None,
        'coverage_end': None,
    }, {}, '/not-applicable.docx')
    ledger_store.insert_payment_plan(contract_id, {
        'phase_name': '年度服务款',
        'due_date': '2026-05-18',
        'due_amount': 86_400,
        'confirm_status': 'confirmed',
    })

    report = ledger_store.build_monthly_payment_report('2026-05')
    assert len(report['rows']) == 1
    assert report['rows'][0]['coverage_not_applicable'] is True
    assert report['diagnostics']['unassigned_serial_count'] == 0
    assert report['diagnostics']['missing_serial_amount_count'] == 0

    output = tmp_path / 'not-applicable-monthly.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)
    workbook = load_workbook(output, data_only=False)
    try:
        detail = next(
            sheet for sheet in workbook.worksheets if sheet.title != '汇总'
        )
        assert detail['A4'].value == '不适用'
    finally:
        workbook.close()


def test_not_applicable_monthly_export_marks_preserved_serial_as_historical(
    tmp_db, tmp_path,
):
    contract_id = ledger_store.create_contract({
        'contract_no': 'HT-NA-HISTORY',
        'title': '历史关联服务合同',
        'project_name': '综合保障项目',
    }, {}, '/not-applicable-history.docx')
    with ledger_store.get_conn() as conn:
        serial_id = conn.execute(
            """
            INSERT INTO contract_serials
                (contract_id, serial_no, status, created_at, updated_at)
            VALUES (?, 9, 'active', ?, ?)
            """,
            (contract_id, '2026-01-01', '2026-01-01'),
        ).lastrowid
    ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serial_id,
        'phase_name': '历史节点',
        'due_date': '2026-05-18',
        'due_amount': 86_400,
        'confirm_status': 'confirmed',
    })
    ledger_store.update_contract(contract_id, {
        'coverage_not_applicable': 1,
        'coverage_start': None,
        'coverage_end': None,
    })

    report = ledger_store.build_monthly_payment_report('2026-05')
    assert report['rows'][0]['coverage_not_applicable'] is True
    assert report['rows'][0]['serial_no'] == 9
    assert report['diagnostics']['unassigned_serial_count'] == 0
    assert report['diagnostics']['missing_serial_amount_count'] == 0

    output = tmp_path / 'not-applicable-history-monthly.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)
    workbook = load_workbook(output, data_only=False)
    try:
        detail = next(
            sheet for sheet in workbook.worksheets if sheet.title != '汇总'
        )
        assert detail['A4'].value == '第 9 发（历史关联；合同发次不适用）'
    finally:
        workbook.close()


def test_payment_plan_inherits_and_overrides_subsystem_and_monthly_splits(
    tmp_db, tmp_path,
):
    contract_id = ledger_store.create_contract({
        'contract_no': 'HT-SUBSYSTEM',
        'title': '分系统归集合同',
        'counterparty': '供应商甲',
        'status': 'active',
        'project_name': '力箭型号',
        'subsystem_name': '总体分系统',
        'coverage_start': 3,
        'coverage_end': 3,
    }, {}, '/subsystem.docx')
    serial = ledger_store.list_contract_serials(contract_id)[0]
    inherited_id = ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serial['id'],
        'phase_name': '总体付款',
        'due_date': '2026-05-10',
        'due_amount': 100_000,
        'confirm_status': 'confirmed',
    })
    override_id = ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serial['id'],
        'subsystem_name': '动力分系统',
        'phase_name': '动力付款',
        'due_date': '2026-05-20',
        'due_amount': 200_000,
        'confirm_status': 'confirmed',
    })

    inherited = ledger_store.get_payment_plan(inherited_id)
    assert inherited['subsystem_name'] == ''
    assert inherited['contract_subsystem_name'] == '总体分系统'
    assert ledger_store.get_payment_plan(override_id)['subsystem_name'] == '动力分系统'

    report = ledger_store.build_monthly_payment_report('2026-05')
    assert len(report['rows']) == 2
    assert {row['subsystem_name'] for row in report['rows']} == {
        '总体分系统', '动力分系统',
    }
    assert {item['subsystem_name'] for item in report['projects']} == {
        '总体分系统', '动力分系统',
    }

    output = tmp_path / 'subsystems.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)
    workbook = load_workbook(output, data_only=False)
    try:
        assert set(workbook.sheetnames) == {
            '汇总', '力箭型号-总体分系统', '力箭型号-动力分系统',
        }
        summary = workbook['汇总']
        assert {
            (summary.cell(row, 2).value, summary.cell(row, 3).value)
            for row in (3, 4)
        } == {
            ('力箭型号', '总体分系统'),
            ('力箭型号', '动力分系统'),
        }
    finally:
        workbook.close()


def test_monthly_report_excludes_void_and_ignores_future_metadata(tmp_db):
    contract_id, serials = _contract_with_serials(title='节点口径合同')
    ledger_store.insert_payment_plans(contract_id, [
        {
            'contract_serial_id': serials[0]['id'],
            'phase_name': '上月未付节点',
            'due_date': '2026-04-20',
            'due_amount': 50_000,
            'confirm_status': 'confirmed',
        },
        {
            'contract_serial_id': serials[0]['id'],
            'phase_name': '本月待确认节点',
            'due_date': '2026-05-20',
            'due_amount': 100_000,
            'confirm_status': 'pending',
        },
        {
            'contract_serial_id': serials[0]['id'],
            'phase_name': '未来尾款',
            'due_date': '2026-08-20',
            'due_amount': 70_000,
            'confirm_status': 'confirmed',
            'remark': '银承：12个月；上月未付原因：未来节点说明',
        },
        {
            'contract_serial_id': serials[0]['id'],
            'phase_name': '作废节点',
            'due_date': '2026-05-25',
            'due_amount': 900_000,
            'confirm_status': 'void',
            'remark': '银承：作废节点；上月未付原因：作废节点说明',
        },
    ])

    report = ledger_store.build_monthly_payment_report('2026-05')
    assert report['node_count'] == 3
    assert len(report['rows']) == 1
    row = report['rows'][0]
    assert row['current_month_minor'] == 10_000_000
    assert row['previous_unpaid_minor'] == 5_000_000
    assert row['planned_payment_minor'] == 15_000_000
    assert row['bank_acceptance'] == ''
    assert row['bank_acceptance_minor'] == 0
    assert row['prior_unpaid_reason'] == ''
    conditions = [node['condition'] for node in row['nodes']]
    assert any('本月待确认节点' in value for value in conditions)
    assert any('未来尾款' in value for value in conditions)
    assert all('作废节点' not in value for value in conditions)


def test_minor_to_wan_uses_round_half_up():
    assert xlsx_exporter._minor_to_wan(None) is None
    assert xlsx_exporter._minor_to_wan(1_005_000) == 1.01
    assert xlsx_exporter._minor_to_wan(2_675_000) == 2.68


def test_monthly_report_export_matches_template_semantics(tmp_db, tmp_path):
    contract_id, serials = _contract_with_serials()
    ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serials[0]['id'],
        'phase_name': '本月进度款',
        'due_date': '2026-05-20',
        'due_amount': 500_000,
        'confirm_status': 'confirmed',
        'condition_text': '完成阶段验收',
    })
    ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serials[0]['id'],
        'phase_name': '未来尾款',
        'due_date': '2026-08-20',
        'due_amount': 300_000,
        'confirm_status': 'confirmed',
        'condition_text': '完成最终验收',
    })
    report = ledger_store.build_monthly_payment_report('2026-05')
    output = tmp_path / 'monthly.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == ['汇总', '火箭项目-未填写分系统']
        summary = workbook['汇总']
        detail = workbook['火箭项目-未填写分系统']
        assert summary['A1'].value is None
        assert summary['B2'].value == '项目'
        assert summary['C2'].value == '所属分系统'
        assert summary['D2'].value == '5月计划付款合计'
        assert summary['E2'].value == '本月计划付款'
        assert summary['F2'].value == '上月已做计划未付款'
        assert summary['G2'].value == '可用银行承兑支付金额'
        assert summary['B3'].value == '火箭项目'
        assert summary['C3'].value == '未填写分系统'
        assert summary['D3'].value == '=SUM(E3:F3)'
        assert summary['B8'].value == '合计：'
        assert summary['B12'].value.startswith('说明：本月计划付款合计包括')
        assert summary.freeze_panes is None
        assert summary.column_dimensions['B'].width == pytest.approx(21.75)
        assert summary['B2'].fill.fgColor.rgb in {'000070C0', '0070C0'}
        assert summary['B2'].font.sz == 14
        assert detail['A3'].value == '火箭发次'
        assert detail['B3'].value == '合同编号'
        assert detail['F3'].value == '合同额'
        assert detail['G3'].value == '付款节点#1\n（金额）'
        assert detail['A4'].value == 101
        assert detail['A4'].number_format == '"第 "0" 发"'
        assert detail['B4'].value == 'HT-月报合同'
        assert detail['F4'].value == 120
        assert detail['F4'].data_type == 'n'
        assert detail['F4'].number_format == '#,##0.00'
        assert detail['F4'].alignment.horizontal == 'right'
        assert detail['G4'].value == 50
        assert detail['G4'].data_type == 'n'
        assert detail['G4'].number_format == '#,##0.00'
        assert detail['G4'].alignment.horizontal == 'right'
        assert detail['G4'].fill.fgColor.rgb in {'00FFFF00', 'FFFF00'}
        assert detail['H4'].fill.fgColor.rgb in {'00FFFF00', 'FFFF00'}
        assert detail['I4'].value == 30
        assert detail['I4'].fill.fill_type is None
        assert detail['K4'].value == '=P4+Q4'
        assert detail['K4'].number_format == '#,##0.00'
        assert detail['K4'].alignment.horizontal == 'right'
        assert detail['K4'].fill.fgColor.rgb in {'00FFFF00', 'FFFF00'}
        assert detail['K11'].value == '=SUM(Q4:Q10)'
        assert detail['K12'].value == '=SUM(P4:P10)'
        assert detail['K13'].value == '=SUM(K11:K12)'
        assert detail['P3'].value is None
        assert detail['Q3'].value is None
        assert detail.column_dimensions['P'].hidden is True
        assert detail.column_dimensions['Q'].hidden is True
        assert detail['M13'].value == '可用银承支付金额'
        assert detail['A15'].value == '绿色的表示已付款'
        assert detail['A16'].value == '黄色表示2026年5月计划付款'
        assert detail.freeze_panes == 'F4'
        assert detail.row_dimensions[1].height == 67
        assert detail.row_dimensions[3].height == 60
        assert str(detail.page_setup.paperSize) == detail.PAPERSIZE_A4
        assert detail.page_setup.orientation == 'portrait'
        assert detail.print_title_rows == '$1:$3'
        assert detail.print_area
        assert summary['E3'].value == "='火箭项目-未填写分系统'!K12"
        assert summary['F3'].value == "='火箭项目-未填写分系统'!K11"
        assert summary['E3'].number_format == '#,##0.00'
        assert summary['E3'].alignment.horizontal == 'right'
        assert summary.print_area == "'汇总'!$B$2:$H$12"
    finally:
        workbook.close()


def test_monthly_export_uses_actual_unique_sheet_titles_and_safe_project_names(
    tmp_db,
    tmp_path,
):
    for index, project_name in enumerate(('Alpha', 'alpha', '=1+1'), 1):
        contract_id, serials = _contract_with_serials(
            title=f'项目名安全合同{index}',
            project=project_name,
        )
        ledger_store.insert_payment_plan(contract_id, {
            'contract_serial_id': serials[0]['id'],
            'phase_name': '本月付款',
            'due_date': '2026-05-20',
            'due_amount': 100_000,
            'confirm_status': 'confirmed',
        })

    report = ledger_store.build_monthly_payment_report('2026-05')
    output = tmp_path / 'monthly-safe-sheet-names.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == [
            '汇总', '=1+1-未填写分系统',
            'Alpha-未填写分系统', 'alpha-未填写分系统-2',
        ]
        summary = workbook['汇总']
        project_rows = {
            summary.cell(row_index, 2).value: row_index
            for row_index in range(3, 6)
        }
        unsafe_name_cell = summary.cell(project_rows["'=1+1"], 2)
        assert unsafe_name_cell.value == "'=1+1"
        assert unsafe_name_cell.data_type == 's'
        assert summary.cell(project_rows['Alpha'], 5).value.startswith(
            "='Alpha-未填写分系统'!"
        )
        assert summary.cell(project_rows['alpha'], 5).value.startswith(
            "='alpha-未填写分系统-2'!"
        )
        for row_index in range(3, 6):
            formula = summary.cell(row_index, 5).value
            referenced_sheet = formula.split("'!", 1)[0][2:].replace("''", "'")
            assert referenced_sheet in workbook.sheetnames
    finally:
        workbook.close()


def test_monthly_report_expands_rows_for_long_text(tmp_db, tmp_path):
    project_name = '液体运载火箭贮箱焊接生产线研发与制造'
    contract_id, serials = _contract_with_serials(
        title='超长合同名称用于验证导出工作表自动增加行高',
        project=project_name,
    )
    ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serials[0]['id'],
        'phase_name': '本月长条件付款',
        'due_date': '2026-05-20',
        'due_amount': 500_000,
        'confirm_status': 'confirmed',
        'condition_text': '完成全部阶段验收并提交完整签字盖章资料',
    })
    report = ledger_store.build_monthly_payment_report('2026-05')
    output = tmp_path / 'monthly-long-text.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)

    workbook = load_workbook(output, data_only=False)
    try:
        summary = workbook['汇总']
        detail = workbook[workbook.sheetnames[1]]
        assert summary['B3'].alignment.wrap_text is True
        assert summary.row_dimensions[3].height > 20
        assert detail.row_dimensions[4].height > 33
    finally:
        workbook.close()


def test_monthly_report_route_and_payment_page(client):
    contract_id, serials = _contract_with_serials()
    ledger_store.insert_payment_plan(contract_id, {
        'contract_serial_id': serials[0]['id'],
        'phase_name': '本月付款',
        'due_date': '2026-05-20',
        'due_amount': 100_000,
        'confirm_status': 'confirmed',
    })

    page = client.get('/payment-plans')
    assert page.status_code == 200
    page_html = page.get_data(as_text=True)
    assert '导出付款计划' in page_html
    assert '导出月度模板' not in page_html
    assert 'type="month"' in page_html
    detail = client.get(f'/contracts/{contract_id}?tab=payments')
    detail_html = detail.get_data(as_text=True)
    assert '发次台账' in detail_html
    assert 'name="plan_0_contract_serial_id"' in detail_html
    assert '不依赖投产通知' in detail_html

    with client.session_transaction() as flask_session:
        flask_session['_csrf_token'] = 'monthly-report-token'
    response = client.post(
        '/payment-plans/export',
        data={
            'csrf_token': 'monthly-report-token',
            'report_month': '2026-05',
        },
    )
    assert response.status_code == 200
    assert response.mimetype == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    assert (
        '2026%E5%B9%B45%E6%9C%88%E5%90%88%E5%90%8C%E4%BB%98%E6%AC%BE'
        '%E8%AE%A1%E5%88%92.xlsx'
    ) in response.headers['Content-Disposition']
    workbook = load_workbook(io.BytesIO(response.data), data_only=False)
    try:
        assert workbook.sheetnames == ['汇总', '火箭项目-未填写分系统']
    finally:
        workbook.close()
    assert client.post(
        '/payment-plans/export',
        data={
            'csrf_token': 'monthly-report-token',
            'report_month': '2026-13',
        },
    ).status_code == 400


def test_empty_monthly_report_has_no_circular_total(tmp_db, tmp_path):
    report = ledger_store.build_monthly_payment_report('2026-05')
    output = tmp_path / 'empty-monthly.xlsx'
    xlsx_exporter.export_monthly_payment_plan_report(output, report)
    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == ['汇总']
        assert workbook['汇总']['B8'].value == '合计：'
        assert workbook['汇总']['D8'].value == '=SUM(E8:F8)'
        assert workbook['汇总']['E8'].value == 0
        assert workbook['汇总']['F8'].value == 0
        assert workbook['汇总']['G8'].value == 0
        assert workbook['汇总']['B12'].value.startswith(
            '说明：本月计划付款合计包括'
        )
    finally:
        workbook.close()
