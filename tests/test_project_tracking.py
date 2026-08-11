import ledger_store
import xlsx_exporter
from openpyxl import load_workbook
from utils.generation_utils import parse_contract_classification


def test_project_classification_validation():
    assert parse_contract_classification({
        'project_name': ' 东区改造 ',
        'subsystem_name': ' 动力分系统 ',
        'coverage_mode': 'range',
        'coverage_start': '11',
        'coverage_end': '28',
    }) == {
        'project_name': '东区改造',
        'subsystem_name': '动力分系统',
        'coverage_mode': 'range',
        'coverage_not_applicable': False,
        'coverage_start': 11,
        'coverage_end': 28,
    }

    for form in (
        {'project_name': '东区改造', 'coverage_start': '11'},
        {
            'project_name': '东区改造', 'coverage_mode': 'range',
            'coverage_start': '11',
        },
        {
            'coverage_mode': 'range', 'coverage_start': '11',
            'coverage_end': '28',
        },
        {
            'project_name': '东区改造', 'coverage_mode': 'range',
            'coverage_start': '28', 'coverage_end': '11',
        },
        {
            'coverage_mode': 'not_applicable', 'coverage_start': '1',
        },
    ):
        try:
            parse_contract_classification(form)
        except ValueError:
            pass
        else:
            raise AssertionError('invalid classification should be rejected')


def test_project_progress_and_payment_association(tmp_db):
    first_id = ledger_store.create_contract({
        'title': '首段合同',
        'status': 'signed',
        'project_name': '东区改造',
        'coverage_start': 1,
        'coverage_end': 20,
    }, {}, '/first.docx')
    second_id = ledger_store.create_contract({
        'title': '二段合同',
        'status': 'active',
        'project_name': '东区改造',
        'coverage_start': 21,
        'coverage_end': 40,
    }, {}, '/second.docx')
    ledger_store.create_contract({
        'title': '待签合同',
        'status': 'draft',
        'project_name': '东区改造',
        'coverage_start': 41,
        'coverage_end': 50,
    }, {}, '/draft.docx')

    ledger_store.insert_payment_plan(first_id, {
        'phase_name': '首付款',
        'confirm_status': 'confirmed',
        'payment_status': 'partial',
        'due_amount': 100,
        'paid_amount': 30,
        'paid_date': '2026-06-01',
    })
    ledger_store.insert_payment_plan(second_id, {
        'phase_name': '进度款',
        'confirm_status': 'pending',
        'payment_status': 'unpaid',
        'due_amount': 200,
    })

    stats = ledger_store.get_project_progress_stats()
    assert len(stats) == 1
    assert stats[0]['project_name'] == '东区改造'
    assert stats[0]['signed_from'] == 1
    assert stats[0]['signed_to'] == 40
    assert stats[0]['planned_to'] == 40
    assert stats[0]['paid_to'] == 20
    assert stats[0]['contract_count'] == 3

    plans = ledger_store.list_payment_plans(project_name='东区改造')
    assert len(plans) == 2
    assert plans[0]['project_name'] == '东区改造'
    assert plans[0]['coverage_start'] == 1
    assert plans[0]['coverage_end'] == 20


def test_project_progress_distinguishes_not_applicable_and_pending(tmp_db):
    not_applicable_id = ledger_store.create_contract({
        'title': '无需发次的已签合同',
        'status': 'signed',
        'project_name': '状态区分项目',
        'coverage_mode': 'not_applicable',
    }, {}, '/not-applicable.docx')
    pending_id = ledger_store.create_contract({
        'title': '历史待补合同',
        'status': 'signed',
        'project_name': '状态区分项目',
    }, {}, '/pending.docx')
    ledger_store.insert_payment_plan(not_applicable_id, {
        'phase_name': '服务费',
        'confirm_status': 'confirmed',
        'payment_status': 'paid',
        'due_amount': 100,
        'paid_amount': 100,
        'paid_date': '2026-05-18',
    })
    ledger_store.insert_payment_plan(pending_id, {
        'phase_name': '待补节点',
        'confirm_status': 'pending',
        'due_amount': 50,
    })

    stats = ledger_store.get_project_progress_stats()[0]
    assert stats['signed_contract_count'] == 2
    assert stats['signed_from'] is None
    assert stats['signed_to'] is None
    assert stats['signed_not_applicable_count'] == 1
    assert stats['signed_pending_coverage_count'] == 1
    assert stats['planned_contract_count'] == 2
    assert stats['planned_not_applicable_count'] == 1
    assert stats['planned_pending_coverage_count'] == 1
    assert stats['paid_contract_count'] == 1
    assert stats['paid_not_applicable_count'] == 1
    assert stats['paid_pending_coverage_count'] == 0


def test_project_fields_can_be_updated_and_searched(tmp_db):
    contract_id = ledger_store.create_contract({'title': '归类测试'}, {}, '/test.docx')
    ledger_store.update_contract(contract_id, {
        'project_name': '西区扩容',
        'subsystem_name': '控制分系统',
        'coverage_start': 7,
        'coverage_end': 19,
    })

    contract = ledger_store.get_contract(contract_id)
    assert contract['project_name'] == '西区扩容'
    assert contract['subsystem_name'] == '控制分系统'
    assert contract['coverage_start'] == 7
    assert contract['coverage_end'] == 19
    assert ledger_store.list_project_names() == ['西区扩容']
    assert ledger_store.list_contracts(q='西区扩容')['total'] == 1
    assert ledger_store.list_contracts(q='控制分系统')['total'] == 1
    subsystem_history = [
        row for row in ledger_store.get_contract_history(contract_id)
        if row['field'] == 'subsystem_name'
    ]
    assert subsystem_history[-1]['new_value'] == '控制分系统'


def test_project_fields_are_in_xlsx_exports(tmp_path):
    payment_path = tmp_path / 'payments.xlsx'
    xlsx_exporter.export_payment_plans(str(payment_path), [{
        'project_name': '东区改造',
        'subsystem_name': '动力分系统',
        'serial_no': 8,
        'coverage_start': 1,
        'coverage_end': 36,
        'contract_no': 'HT-01',
        'contract_title': '测试合同',
    }])
    payment_sheet = load_workbook(payment_path).active
    assert payment_sheet['B3'].value == '项目名称'
    assert payment_sheet['C3'].value == '所属分系统'
    assert payment_sheet['D3'].value == '所属发次'
    assert payment_sheet['E3'].value == '发次范围'
    assert payment_sheet['B4'].value == '东区改造'
    assert payment_sheet['C4'].value == '动力分系统'
    assert payment_sheet['D4'].value == '第 8 发'
    assert payment_sheet['E4'].value == '第 1–36 发'

    contract_path = tmp_path / 'contracts.xlsx'
    xlsx_exporter.export_contracts(str(contract_path), [{
        'project_name': '东区改造',
        'subsystem_name': '动力分系统',
        'coverage_start': 1,
        'coverage_end': 36,
        'title': '测试合同',
    }])
    contract_sheet = load_workbook(contract_path).active
    assert contract_sheet['B3'].value == '项目名称'
    assert contract_sheet['C3'].value == '所属分系统'
    assert contract_sheet['D3'].value == '发次范围'
    assert contract_sheet['B4'].value == '东区改造'
    assert contract_sheet['C4'].value == '动力分系统'
    assert contract_sheet['D4'].value == '第 1–36 发'


def test_pending_coverage_is_labeled_in_xlsx_exports(tmp_path):
    payment_path = tmp_path / 'pending-payments.xlsx'
    xlsx_exporter.export_payment_plans(str(payment_path), [{
        'contract_title': '历史待补合同',
    }])
    payment_sheet = load_workbook(payment_path).active
    assert payment_sheet['D4'].value == '待补发次'
    assert payment_sheet['E4'].value == '待补发次'

    for streaming in (False, True):
        contract_path = tmp_path / f'pending-contracts-{streaming}.xlsx'
        xlsx_exporter.export_contracts(
            str(contract_path),
            [{'title': '历史待补合同'}],
            streaming=streaming,
        )
        contract_sheet = load_workbook(contract_path).active
        assert contract_sheet['D4'].value == '待补发次'
