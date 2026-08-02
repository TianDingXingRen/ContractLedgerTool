"""Central style factory for generated workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class WorkbookStyleFactory:
    """Create and apply the small, shared workbook design system."""

    money_format = '#,##0.00'

    def __init__(self):
        self.header_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=11,
        )
        self.title_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=16,
        )
        self.normal_font = Font(name='Microsoft YaHei', size=11)
        self.note_font = Font(
            name='Microsoft YaHei',
            size=10,
            color='666666',
        )
        self.primary_header_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=10,
            color='FFFFFF',
        )
        self.primary_summary_header_font = Font(
            name='Microsoft YaHei',
            bold=True,
            color='FFFFFF',
        )
        self.monthly_summary_header_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=14,
            color='FFFFFF',
        )
        self.monthly_header_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=14,
            color='FFFFFF',
        )
        self.monthly_title_font = Font(
            name='Microsoft YaHei',
            size=24,
        )
        self.monthly_note_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=16,
        )
        self.monthly_body_font = Font(name='Microsoft YaHei', size=11)
        self.monthly_summary_body_font = Font(name='Microsoft YaHei', size=14)
        self.monthly_summary_bold_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=14,
        )
        self.monthly_previous_font = Font(
            name='Microsoft YaHei',
            size=11,
            color='FF0000',
        )
        self.monthly_total_font = Font(
            name='Microsoft YaHei',
            bold=True,
            size=11,
        )

        self.header_fill = self.solid_fill('D9E1F2')
        self.primary_header_fill = self.solid_fill('4472C4')
        self.paid_fill = self.solid_fill('C6E0B4')
        self.current_fill = self.solid_fill('FFF2CC')
        self.warning_fill = self.solid_fill('FCE4D6')
        self.monthly_header_fill = self.solid_fill('0070C0')
        self.monthly_paid_fill = self.solid_fill('00B050')
        self.monthly_current_fill = self.solid_fill('FFFF00')
        self.monthly_helper_fill = self.solid_fill('DDEBF7')
        self.monthly_total_fill = self.solid_fill('F4CCCC')

        thin = Side(style='thin')
        self.thin_border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        self.title_alignment = Alignment(horizontal='center')
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
        )
        self.center_wrap_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )
        self.note_alignment = Alignment(
            vertical='center',
            wrap_text=True,
        )
        self.top_wrap_alignment = Alignment(
            vertical='top',
            wrap_text=True,
        )
        self.left_wrap_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True,
        )
        self.right_wrap_alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=True,
        )
        self.right_alignment = Alignment(horizontal='right')

    @staticmethod
    def solid_fill(color):
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid',
        )

    def apply_header_row(
        self,
        worksheet,
        row,
        column_count,
        *,
        primary=False,
        compact=False,
    ):
        font = self.header_font
        fill = self.header_fill
        alignment = self.center_alignment
        if primary:
            font = (
                self.primary_header_font
                if compact
                else self.primary_summary_header_font
            )
            fill = self.primary_header_fill
            alignment = self.center_wrap_alignment
        for column in range(1, column_count + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = self.thin_border

    @staticmethod
    def apply_column_widths(worksheet, columns):
        for index, column in enumerate(columns, start=1):
            worksheet.column_dimensions[
                get_column_letter(index)
            ].width = column.width
